"""
sync_tags.py
------------
1. Tạo Tags (res.partner.category) từ file nhóm MISA
2. Gán đúng tag cho từng KH/NCC dựa trên mã nhóm đã lưu trong ghi chú
"""
import openpyxl, xmlrpc.client, sys, re
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nhom_khach_hang_nha_cung_cap.xlsx"
ODOO_URL   = 'http://localhost:8017'
ODOO_DB    = 'odoo_company'
ODOO_USER  = 'doanvanky36k21@gmail.com'
ODOO_PASS  = 'admin'

def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{ts}] {msg}")

# ---- Màu tag (xoay vòng 10 màu Odoo) ----
COLORS = [1,2,3,4,5,6,7,8,9,10,11]

# ---- Kết nối ----
common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
log("Kết nối Odoo OK")

# ---- Đọc file nhóm ----
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active
groups = []
for r in range(4, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 6)]
    stt, ma_nhom, ten_nhom, dien_giai, trang_thai = row
    if not ma_nhom or not ten_nhom: continue
    groups.append({
        'ma':       str(ma_nhom).strip(),
        'ten':      str(ten_nhom).strip(),
        'mo_ta':    str(dien_giai).strip() if dien_giai else '',
    })
log(f"Đọc được {len(groups)} nhóm từ Excel")

# ---- Tạo/lấy tags ----
existing_tags = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'res.partner.category', 'search_read', [[]], {'fields': ['id','name']})
tag_by_name = {t['name']: t['id'] for t in existing_tags}
tag_by_ma   = {}  # ma_nhom -> tag_id

for i, grp in enumerate(groups):
    # Tên tag: "CG - Chủ gỗ"
    tag_name = f"{grp['ma']} - {grp['ten']}"
    if tag_name in tag_by_name:
        tag_id = tag_by_name[tag_name]
        log(f"  ↻ Tag đã có: {tag_name}")
    else:
        tag_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
            'res.partner.category', 'create', [{
                'name':  tag_name,
                'color': COLORS[i % len(COLORS)],
            }])
        tag_by_name[tag_name] = tag_id
        log(f"  ✓ Tạo tag: {tag_name} (id={tag_id})")
    tag_by_ma[grp['ma']] = tag_id

log(f"\nĐã xử lý {len(tag_by_ma)} tags")

# ---- Gán tags cho partners ----
# Đọc tất cả partners có comment chứa "MISA"
log("\nĐang gán tags cho KH/NCC...")
partners = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'res.partner', 'search_read',
    [[['comment', 'like', 'MISA']]],
    {'fields': ['id', 'name', 'comment', 'category_id']})

log(f"Tìm thấy {len(partners)} partners có dữ liệu MISA")

assigned = 0
for p in partners:
    comment = p.get('comment') or ''
    # Tìm "Nhóm=XX" trong comment
    match = re.search(r'Nh[oó]m=([A-Z0-9]+)', comment)
    if not match:
        continue
    ma_nhom = match.group(1).strip()
    if ma_nhom not in tag_by_ma:
        continue

    tag_id = tag_by_ma[ma_nhom]
    current_tags = [t['id'] if isinstance(t, dict) else t for t in p.get('category_id', [])]

    if tag_id not in current_tags:
        new_tags = current_tags + [tag_id]
        models.execute_kw(ODOO_DB, uid, ODOO_PASS,
            'res.partner', 'write',
            [[p['id']], {'category_id': [(6, 0, new_tags)]}])
        assigned += 1

log(f"\nKẾT QUẢ:")
log(f"  Tags tạo: {len(tag_by_ma)}")
log(f"  Partners được gán tag: {assigned}")
log("HOÀN TẤT")
