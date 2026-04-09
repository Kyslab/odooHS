"""
sync_ma_nha_cung_cap.py
Đọc Danh_sach_nha_cung_cap.xlsx → gán Mã NCC (ref) cho supplier trong Odoo
Khớp theo thứ tự: ref sẵn → SĐT → tên chuẩn hóa → tạo mới
"""
import openpyxl, psycopg2, re, json, sys, logging
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(filename=r'D:\odoo\sync_ma_nha_cung_cap.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8')

EXCEL_PATH = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nha_cung_cap.xlsx"

def log(msg): print(msg); logging.info(msg)

def norm_phone(p):
    if not p: return ''
    s = re.sub(r'[^0-9]', '', str(p))
    if s.startswith('84') and len(s) >= 11: s = '0' + s[2:]
    return s

def norm_name(s):
    s = str(s).lower().strip()
    s = re.sub(r'\b0[0-9]{9,10}\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ─── 1. Đọc Excel ─────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

file_suppliers = []
for r in range(4, ws.max_row + 1):
    ma_ncc  = ws.cell(row=r, column=2).value
    ten_ncc = ws.cell(row=r, column=3).value
    dia_chi = ws.cell(row=r, column=4).value
    mst     = ws.cell(row=r, column=7).value
    phone   = ws.cell(row=r, column=8).value
    is_cust = ws.cell(row=r, column=10).value

    if not ma_ncc or not ten_ncc: continue
    ma_ncc  = str(ma_ncc).strip()
    ten_ncc = str(ten_ncc).strip()
    if ten_ncc in ('#N/A', 'Tổng', '') or ma_ncc in ('0', '#N/A', 'Tổng'):
        continue
    # Chấp nhận tất cả mã (kể cả không theo chuẩn như BIDV, KDTV, anh Dâng)

    file_suppliers.append({
        'ma_ncc'  : ma_ncc,
        'ten_ncc' : ten_ncc,
        'dia_chi' : str(dia_chi).strip() if dia_chi else '',
        'mst'     : str(mst).strip() if mst else '',
        'phone'   : norm_phone(phone),
        'is_cust' : bool(is_cust and str(is_cust).strip() in ('✓','x','X','1')),
    })

log(f"  {len(file_suppliers)} NCC hợp lệ từ file")

# ─── 2. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()

cur.execute("SELECT id, name, ref, phone, mobile, vat FROM res_partner WHERE active=TRUE;")
ref_map = {}; phone_map = {}; name_map = {}
for pid, pname, pref, pphone, pmobile, pvat in cur.fetchall():
    nm = list(pname.values())[0] if isinstance(pname, dict) else str(pname) if pname else ''
    if pref: ref_map[pref.strip()] = pid
    for ph in [pphone, pmobile]:
        np = norm_phone(ph)
        if np and len(np) >= 9: phone_map.setdefault(np, pid)
    nn = norm_name(nm)
    if nn: name_map.setdefault(nn, pid)

log(f"  {len(ref_map)} ref | {len(phone_map)} phone | {len(name_map)} name trong Odoo")

# ─── 3. Khớp và cập nhật ─────────────────────────────────────────────────────
log("\n🔗 Khớp Mã NCC...")
stats = {'ref': 0, 'phone': 0, 'name': 0, 'new': 0}
not_matched = []

for c in file_suppliers:
    ma = c['ma_ncc']

    # 1. Đã có ref
    if ma in ref_map:
        # Cập nhật thêm thông tin còn thiếu
        cur.execute("""
            UPDATE res_partner SET
                street       = CASE WHEN (street IS NULL OR street='') AND %s!='' THEN %s ELSE street END,
                vat          = CASE WHEN (vat IS NULL OR vat='') AND %s!='' THEN %s ELSE vat END,
                phone        = CASE WHEN (phone IS NULL OR phone='') AND %s!='' THEN %s ELSE phone END,
                supplier_rank= GREATEST(supplier_rank, 1),
                customer_rank= GREATEST(customer_rank, %s),
                write_date=NOW(), write_uid=2
            WHERE id=%s;
        """, (c['dia_chi'],c['dia_chi'], c['mst'],c['mst'],
              c['phone'],c['phone'], 1 if c['is_cust'] else 0, ref_map[ma]))
        stats['ref'] += 1
        continue

    # 2. Khớp SĐT
    found_id = None
    if c['phone'] and c['phone'] in phone_map:
        found_id = phone_map[c['phone']]

    if found_id:
        cur.execute("""
            UPDATE res_partner SET ref=%s, supplier_rank=GREATEST(supplier_rank,1),
                write_date=NOW(), write_uid=2 WHERE id=%s AND (ref IS NULL OR ref='');
        """, (ma, found_id))
        ref_map[ma] = found_id
        stats['phone'] += 1
        continue

    # 3. Khớp tên
    nn = norm_name(c['ten_ncc'])
    if nn in name_map:
        found_id = name_map[nn]
        cur.execute("""
            UPDATE res_partner SET ref=%s, supplier_rank=GREATEST(supplier_rank,1),
                write_date=NOW(), write_uid=2 WHERE id=%s AND (ref IS NULL OR ref='');
        """, (ma, found_id))
        ref_map[ma] = found_id
        stats['name'] += 1
        continue

    # 4. Không tìm thấy → tạo mới
    not_matched.append(c)

log(f"  Đã có ref: {stats['ref']} | SĐT: {stats['phone']} | Tên: {stats['name']} | Cần tạo mới: {len(not_matched)}")

# ─── 4. Tạo mới NCC chưa có ──────────────────────────────────────────────────
if not_matched:
    log(f"\n➕ Tạo mới {len(not_matched)} NCC...")
    for c in not_matched:
        name_json = json.dumps({"vi_VN": c['ten_ncc'], "en_US": c['ten_ncc']}, ensure_ascii=False)
        cur.execute("""
            INSERT INTO res_partner
                (name, ref, street, vat, phone, supplier_rank, customer_rank,
                 active, is_company, create_uid, write_uid, create_date, write_date)
            VALUES (%s,%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),
                    1, %s, TRUE, FALSE, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (name_json, c['ma_ncc'], c['dia_chi'], c['mst'], c['phone'] or None,
              1 if c['is_cust'] else 0))
        new_id = cur.fetchone()[0]
        ref_map[c['ma_ncc']] = new_id
        stats['new'] += 1
        log(f"  ➕ {c['ma_ncc']} - {c['ten_ncc']}")

conn.commit()
cur.close(); conn.close()

log(f"""
╔══════════════════════════════════════════════════╗
║       HOÀN TẤT GÁN MÃ NHÀ CUNG CẤP             ║
╠══════════════════════════════════════════════════╣
║  Tổng NCC trong file  : {len(file_suppliers):<24} ║
║  Đã có ref sẵn        : {stats['ref']:<24} ║
║  Khớp qua SĐT         : {stats['phone']:<24} ║
║  Khớp qua tên         : {stats['name']:<24} ║
║  Tạo mới              : {stats['new']:<24} ║
╚══════════════════════════════════════════════════╝
""")
