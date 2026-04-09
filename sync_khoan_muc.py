"""
sync_khoan_muc.py
Nhập 706 khoản mục chi phí từ MISA vào Odoo 17 (analytic plan riêng)
+ Bắt buộc chọn khoản mục khi nhập journal entry tài khoản 6xx
"""

import openpyxl, psycopg2, json, re, sys, logging
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    filename=r'D:\odoo\sync_khoan_muc.log',
    filemode='w',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    encoding='utf-8'
)

EXCEL_PATH = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khoan_muc_chi_phi_.xlsx"
PLAN_NAME_VI = "Khoản mục chi phí"
PLAN_NAME_EN = "Expense Categories"
# Bắt buộc chọn khoản mục khi nhập dòng tài khoản bắt đầu bằng '6' hoặc '7'
ACCOUNT_PREFIXES_MANDATORY = ['6', '7']

def log(msg):
    print(msg)
    logging.info(msg)

# ─── 1. Đọc Excel ────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

items = []  # list of {'code': str, 'name': str}
for r in range(4, ws.max_row + 1):
    code  = ws.cell(row=r, column=2).value
    name  = ws.cell(row=r, column=3).value
    state = ws.cell(row=r, column=5).value
    if not code or not name:
        continue
    code  = str(code).strip()
    name  = str(name).strip()
    state = str(state).strip() if state else ''
    if state == 'Ngừng sử dụng':
        continue
    items.append({'code': code, 'name': name})

log(f"  → Đọc được {len(items)} khoản mục đang sử dụng")

# ─── 2. Xây dựng bảng tra cứu parent theo prefix ─────────────────────────────
all_codes = {item['code'] for item in items}

def find_parent_code(code):
    """Tìm code cha = code con dài nhất là prefix của code hiện tại"""
    for length in range(len(code) - 1, 0, -1):
        prefix = code[:length]
        if prefix in all_codes:
            return prefix
    return None

# Xây bảng parent_code
parent_map = {}  # code → parent_code (or None)
for item in items:
    parent_map[item['code']] = find_parent_code(item['code'])

# ─── 3. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(
    host='localhost', port=5432, dbname='odoo_company',
    user='odoo17', password='odoo17pass'
)
conn.autocommit = False
cur = conn.cursor()

try:
    # ─── 4. Lấy company_id ───────────────────────────────────────────────────
    cur.execute("SELECT id FROM res_company ORDER BY id LIMIT 1;")
    company_id = cur.fetchone()[0]
    log(f"  Company ID: {company_id}")

    # ─── 5. Tạo hoặc lấy Analytic Plan "Khoản mục chi phí" qua XML-RPC ──────
    # QUAN TRỌNG: phải tạo plan qua ORM (XML-RPC), không dùng SQL trực tiếp
    # vì Odoo cần tự tạo dynamic field x_planN_id trên account_analytic_line
    import xmlrpc.client
    ODOO_URL  = 'http://localhost:8017'
    ODOO_USER = 'doanvanky36k21@gmail.com'
    ODOO_PASS = 'admin'

    xmodels = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')

    cur.execute("SELECT id FROM account_analytic_plan WHERE name::text ILIKE %s;",
                (f'%{PLAN_NAME_VI}%',))
    row = cur.fetchone()
    if row:
        plan_id = row[0]
        log(f"  Plan đã tồn tại: ID={plan_id}")
    else:
        # Tạo qua XML-RPC để Odoo sinh field động x_planN_id
        plan_id = xmodels.execute_kw(
            'odoo_company', 2, ODOO_PASS,
            'account.analytic.plan', 'create',
            [{'name': PLAN_NAME_VI, 'description': PLAN_NAME_EN}]
        )
        log(f"  ✅ Đã tạo plan mới qua ORM: ID={plan_id} — {PLAN_NAME_VI}")

    # ─── 6. Lấy danh sách khoản mục đã có trong DB ───────────────────────────
    cur.execute("""
        SELECT code FROM account_analytic_account
        WHERE plan_id = %s;
    """, (plan_id,))
    existing_codes = {r[0] for r in cur.fetchall() if r[0]}
    log(f"  Đã có {len(existing_codes)} khoản mục trong DB")

    # ─── 7. Import từng khoản mục — 2 lượt (lượt 1: không có parent; lượt 2: có parent) ──
    code_to_id = {}  # code → account_id (sau khi insert)

    # Lấy các id đã tồn tại vào code_to_id
    if existing_codes:
        cur.execute("""
            SELECT code, id FROM account_analytic_account
            WHERE plan_id = %s AND code IS NOT NULL;
        """, (plan_id,))
        for r in cur.fetchall():
            code_to_id[r[0]] = r[1]

    # Sắp xếp theo độ dài code để insert cha trước con
    items_sorted = sorted(items, key=lambda x: len(x['code']))

    created = 0
    skipped = 0

    for item in items_sorted:
        code = item['code']
        name = item['name']

        if code in existing_codes:
            skipped += 1
            # Cập nhật code_to_id nếu chưa có
            if code not in code_to_id:
                cur.execute("SELECT id FROM account_analytic_account WHERE plan_id=%s AND code=%s;",
                            (plan_id, code))
                r = cur.fetchone()
                if r:
                    code_to_id[code] = r[0]
            continue

        name_json = json.dumps({"vi_VN": name, "en_US": name}, ensure_ascii=False)
        parent_code = parent_map.get(code)
        parent_id = code_to_id.get(parent_code) if parent_code else None

        cur.execute("""
            INSERT INTO account_analytic_account
                (name, code, plan_id, root_plan_id, company_id, active, create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, %s, %s, %s, TRUE, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (name_json, code, plan_id, plan_id, company_id))
        new_id = cur.fetchone()[0]
        code_to_id[code] = new_id
        created += 1

        if created % 100 == 0:
            log(f"    ... đã nhập {created} khoản mục")

    log(f"\n  ✅ Kết quả: Tạo mới {created}, Bỏ qua (đã có) {skipped}")

    # ─── 8. Cập nhật parent_id ───────────────────────────────────────────────
    log("\n  🔗 Cập nhật quan hệ cha-con...")
    updated_parents = 0
    for item in items_sorted:
        code = item['code']
        parent_code = parent_map.get(code)
        if parent_code and code in code_to_id and parent_code in code_to_id:
            cur.execute("""
                UPDATE account_analytic_account
                SET write_date = NOW()
                WHERE id = %s AND plan_id = %s;
            """, (code_to_id[code], plan_id))
            # Note: account_analytic_account in Odoo 17 Community may not have parent_id
            # Check if column exists
    # Actually check if parent_id column exists in account_analytic_account
    cur.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name='account_analytic_account' AND column_name='parent_id';
    """)
    has_parent_col = cur.fetchone() is not None
    log(f"  account_analytic_account.parent_id exists: {has_parent_col}")

    if has_parent_col:
        for item in items_sorted:
            code = item['code']
            parent_code = parent_map.get(code)
            if parent_code and code in code_to_id and parent_code in code_to_id:
                cur.execute("""
                    UPDATE account_analytic_account
                    SET parent_id = %s, write_date = NOW()
                    WHERE id = %s AND plan_id = %s;
                """, (code_to_id[parent_code], code_to_id[code], plan_id))
                updated_parents += 1
        log(f"  ✅ Đã cập nhật {updated_parents} quan hệ cha-con")
    else:
        log("  ℹ️  Không có cột parent_id — dữ liệu phẳng (flat), không phân cấp")

    # ─── 9. Cấu hình bắt buộc chọn khoản mục (account_analytic_applicability) ──
    log("\n  ⚙️  Cấu hình bắt buộc chọn khoản mục chi phí...")
    for prefix in ACCOUNT_PREFIXES_MANDATORY:
        # Kiểm tra đã có chưa
        cur.execute("""
            SELECT id FROM account_analytic_applicability
            WHERE analytic_plan_id = %s
              AND business_domain = 'general'
              AND account_prefix = %s;
        """, (plan_id, prefix))
        existing = cur.fetchone()
        if existing:
            # Cập nhật thành mandatory
            cur.execute("""
                UPDATE account_analytic_applicability
                SET applicability = 'mandatory', write_date = NOW(), write_uid = 2
                WHERE id = %s;
            """, (existing[0],))
            log(f"  ✅ Đã CẬP NHẬT applicability → mandatory cho tài khoản prefix '{prefix}'")
        else:
            cur.execute("""
                INSERT INTO account_analytic_applicability
                    (analytic_plan_id, company_id, business_domain, applicability,
                     account_prefix, create_uid, write_uid, create_date, write_date)
                VALUES (%s, %s, 'general', 'mandatory', %s, 2, 2, NOW(), NOW());
            """, (plan_id, company_id, prefix))
            log(f"  ✅ Đã TẠO quy tắc bắt buộc cho tài khoản prefix '{prefix}'")

    conn.commit()
    log("\n✅ HOÀN TẤT! Đã commit tất cả thay đổi.")
    log(f"   Plan ID: {plan_id}")
    log(f"   Tổng khoản mục đã nhập: {created + skipped}")
    log(f"   Mới tạo: {created}")
    log(f"   Đã có sẵn: {skipped}")

except Exception as e:
    conn.rollback()
    log(f"\n❌ LỖI: {e}")
    import traceback
    traceback.print_exc()
    raise
finally:
    cur.close()
    conn.close()
