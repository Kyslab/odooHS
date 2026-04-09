"""
sync_suppliers.py
-----------------
Import / đồng bộ nhà cung cấp từ MISA → Odoo.
- Nếu NCC đã tồn tại (trùng mã MISA) → cập nhật supplier_rank=1
- Nếu NCC trùng tên với KH đã có → gộp, chỉ cập nhật supplier_rank
- Nếu chưa có → tạo mới
"""

import openpyxl, xmlrpc.client, sys, re
from datetime import datetime
from odoo_helper import load_all_tags, assign_tag_to_partner
sys.stdout.reconfigure(encoding='utf-8')

# ===================== CẤU HÌNH =====================
EXCEL_FILE  = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nha_cung_cap.xlsx"
SHEET_NAME  = 'Danh sách nhà cung cấp'
DATA_START  = 4

ODOO_URL    = 'http://localhost:8017'
ODOO_DB     = 'odoo_company'
ODOO_USER   = 'doanvanky36k21@gmail.com'
ODOO_PASS   = 'admin'
EXT_MODULE  = '__import__'
LOG_FILE    = r"D:\odoo\sync_suppliers.log"
# ====================================================

def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

def clean(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s in ('None', '#N/A', 'None') else s

def make_ext_id(ma):
    return f"misa_{re.sub(r'[^a-zA-Z0-9_]', '_', ma)}"

def detect_company_type(name):
    name_lower = name.lower()
    for kw in ['công ty', 'cty', 'tnhh', 'cổ phần', 'tập đoàn', 'ngân hàng',
               'bệnh viện', 'trường ', 'htx', 'chi nhánh', 'co.,', 'ltd']:
        if kw in name_lower:
            return 'company'
    for ho in ['nguyễn', 'trần', 'lê', 'phạm', 'huỳnh', 'hoàng', 'phan',
               'vũ', 'võ', 'đặng', 'bùi', 'đỗ', 'hồ', 'ngô', 'dương',
               'anh ', 'chị ', 'ông ', 'bà ']:
        if name_lower.startswith(ho):
            return 'person'
    return 'company'

# ---- Kết nối ----
common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
if not uid:
    log("Lỗi đăng nhập Odoo!"); exit(1)
models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
log("Kết nối Odoo OK")

# ---- Lấy ext_id đã có ----
existing_ext = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'ir.model.data', 'search_read',
    [[['module','=',EXT_MODULE],['model','=','res.partner']]],
    {'fields': ['name','res_id']})
ext_map = {r['name']: r['res_id'] for r in existing_ext}
log(f"Odoo có {len(ext_map)} bản ghi MISA sẵn có")

# ---- Load tags ----
log("Đang load danh mục nhóm (tags)...")
tag_cache = load_all_tags(models, uid, ODOO_DB, ODOO_PASS)
log(f"Đã load {len(tag_cache)} nhóm tags")

# ---- Đọc Excel ----
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

suppliers = []
for r in range(DATA_START, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 11)]
    stt, ma_ncc, ten_ncc, dia_chi, cong_no, nhom, mst, phone, tk_no, la_kh = row

    ma_ncc  = clean(ma_ncc)
    ten_ncc = clean(ten_ncc)
    if not ma_ncc or not ten_ncc or ma_ncc == 'Tổng':
        continue

    mst_clean = clean(mst)
    if not re.match(r'^\d{10}(\d{3})?$', mst_clean):
        mst_clean = ''

    suppliers.append({
        'ext_id':        make_ext_id(ma_ncc),
        'ma_ncc':        ma_ncc,
        'name':          ten_ncc,
        'company_type':  detect_company_type(ten_ncc),
        'street':        clean(dia_chi),
        'phone':         clean(phone),
        'vat':           mst_clean,
        'supplier_rank': 1,
        'customer_rank': 1 if clean(la_kh) == '✓' else 0,
        'comment':       f"MISA NCC: Mã={ma_ncc} | Nhóm={clean(nhom)} | Công nợ phải trả={cong_no or 0:,.0f} | TK nợ={clean(tk_no)}",
        'cong_no':       cong_no or 0,
        'tk_no':         clean(tk_no),
    })

log(f"Đọc được {len(suppliers)} NCC từ Excel")

# ---- Xử lý từng NCC ----
log("=" * 50)
created = updated = skipped = 0
errors  = []

for rec in suppliers:
    try:
        ext_id = rec['ext_id']

        if ext_id in ext_map:
            # --- Đã có trong Odoo (từ lần import KH trước) → chỉ update supplier_rank ---
            partner_id = ext_map[ext_id]
            vals = {'supplier_rank': 1}
            if rec['customer_rank'] == 1:
                vals['customer_rank'] = 1
            models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'res.partner', 'write', [[partner_id], vals])
            updated += 1
            log(f"  ↻ Cập nhật NCC (đã có): [{rec['ma_ncc']}] {rec['name']}")

        else:
            # --- Chưa có → tạo mới ---
            vals = {
                'name':          rec['name'],
                'company_type':  rec['company_type'],
                'supplier_rank': 1,
                'customer_rank': rec['customer_rank'],
            }
            if rec['street']:  vals['street']  = rec['street']
            if rec['phone']:   vals['phone']   = rec['phone']
            if rec['vat']:     vals['vat']     = rec['vat']
            vals['comment'] = rec['comment']

            partner_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'res.partner', 'create', [vals])

            # Lưu external ID
            models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'ir.model.data', 'create', [{
                    'module':   EXT_MODULE,
                    'name':     ext_id,
                    'model':    'res.partner',
                    'res_id':   partner_id,
                    'noupdate': True,
                }])
            created += 1
            log(f"  ✓ Tạo mới NCC: [{rec['ma_ncc']}] {rec['name']}")
            # Gán tag nhóm
            nhom = re.search(r'Nh[oó]m=([A-Z0-9]+)', rec.get('comment',''))
            if nhom:
                assign_tag_to_partner(models, uid, ODOO_DB, ODOO_PASS, partner_id, nhom.group(1), tag_cache)

    except Exception as e:
        errors.append(f"[{rec['ma_ncc']}] {rec['name']}: {e}")
        log(f"  ✗ Lỗi [{rec['ma_ncc']}] {rec['name']}: {e}")

log("=" * 50)
log(f"KẾT QUẢ:")
log(f"  Tạo mới NCC : {created}")
log(f"  Cập nhật    : {updated} (đã có từ danh sách KH)")
log(f"  Lỗi         : {len(errors)}")
log("ĐỒNG BỘ NHÀ CUNG CẤP HOÀN TẤT")
log("=" * 50)
