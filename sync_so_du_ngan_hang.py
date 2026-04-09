"""
sync_so_du_ngan_hang.py
Nhập số dư đầu kỳ cho từng tài khoản ngân hàng vào Odoo 17
- Xóa dòng TK 1121 trong bút toán MISC đã tạo (tránh double-count)
- Tạo bút toán riêng trong từng journal ngân hàng
- Offset với TK 4111 (Vốn đầu tư của chủ sở hữu)
"""

import openpyxl, psycopg2, json, sys, logging, xmlrpc.client, re
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    filename=r'D:\odoo\sync_so_du_ngan_hang.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8'
)

EXCEL_PATH   = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nhap_so_du_tai_khoan_ngan_hang.xlsx"
OPENING_DATE = '2024-12-31'
MISC_MOVE_ID = 2       # ID bút toán MISC đã tạo ở bước trước
BALANCE_CODE = '4111'  # TK cân đối

ODOO_URL  = 'http://localhost:8017'
ODOO_DB   = 'odoo_company'
ODOO_UID  = 2
ODOO_PASS = 'admin'

def log(msg):
    print(msg); logging.info(msg)

def normalize_acc(s):
    """Chuẩn hóa số TK: chỉ giữ chữ số, bỏ số 0 đầu"""
    return re.sub(r'[^0-9]', '', str(s)).lstrip('0')

# ─── 1. Đọc Excel ─────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

banks = []
for r in range(4, ws.max_row + 1):
    acc_num = ws.cell(row=r, column=1).value
    bank_nm = ws.cell(row=r, column=2).value
    balance = ws.cell(row=r, column=4).value   # Dư Nợ (cột D)
    if not acc_num or balance is None:
        continue
    acc_num = str(acc_num).strip()
    balance = float(balance)
    banks.append({'acc_num': acc_num, 'bank_name': str(bank_nm).strip() if bank_nm else '', 'balance': balance})

log(f"  Đọc được {len(banks)} tài khoản ngân hàng")
total_balance = sum(b['balance'] for b in banks)
log(f"  Tổng số dư: {total_balance:,.0f}")

# ─── 2. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='localhost', port=5432, dbname=ODOO_DB,
                         user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()

# Lấy tất cả journal ngân hàng + số TK
cur.execute("""
    SELECT aj.id, aj.code, aj.name, rpb.acc_number, aa.id as acc_id
    FROM account_journal aj
    JOIN res_partner_bank rpb ON rpb.id = aj.bank_account_id
    JOIN account_account aa ON aa.id = aj.default_account_id
    WHERE aj.type = 'bank';
""")
journals = []
for r in cur.fetchall():
    journals.append({
        'journal_id': r[0],
        'code'      : r[1],
        'name'      : list(r[2].values())[0] if isinstance(r[2], dict) else str(r[2]),
        'acc_number': str(r[3]).strip(),
        'gl_acc_id' : r[4],   # TK 1121
    })

log(f"\n  Có {len(journals)} journal ngân hàng trong Odoo")

# Lấy account_id của TK 4111
cur.execute("SELECT id FROM account_account WHERE code = %s;", (BALANCE_CODE,))
balance_acc_id = cur.fetchone()[0]

# Lấy account_id của TK 1121
cur.execute("SELECT id FROM account_account WHERE code = '1121';")
acc_1121_id = cur.fetchone()[0]
log(f"  TK 1121 ID={acc_1121_id}, TK 4111 ID={balance_acc_id}")

# ─── 3. Khớp tài khoản file ↔ journal Odoo ───────────────────────────────────
log("\n🔗 Khớp số tài khoản...")
matched = []
unmatched = []

for b in banks:
    norm_file = normalize_acc(b['acc_num'])
    found = None
    for j in journals:
        norm_odoo = normalize_acc(j['acc_number'])
        if norm_file == norm_odoo:
            found = j
            break
        # Fallback: một bên là suffix của bên kia
        if norm_file and norm_odoo and (norm_file.endswith(norm_odoo) or norm_odoo.endswith(norm_file)):
            found = j
            break
    if found:
        matched.append({'bank': b, 'journal': found})
        log(f"  ✅ {found['code']:6} | {b['acc_num']:35} | {b['balance']:>15,.0f}")
    else:
        unmatched.append(b)
        log(f"  ❌ KHÔNG KHỚP: {b['acc_num']:35} | {b['balance']:>15,.0f}")

log(f"\n  Khớp: {len(matched)}, Không khớp: {len(unmatched)}")
if unmatched:
    log("  ⚠️  Tài khoản không khớp:")
    for b in unmatched:
        log(f"       {b['acc_num']} | {b['balance']:,.0f}")

# ─── 4. Chỉnh sửa bút toán MISC: xóa dòng TK 1121 ───────────────────────────
log(f"\n✏️  Chỉnh sửa bút toán MISC (ID={MISC_MOVE_ID}): xóa dòng TK 1121...")

xmodels = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object', allow_none=True)

# Kiểm tra trạng thái, chỉ reset nếu đang posted
cur.execute("SELECT state FROM account_move WHERE id=%s;", (MISC_MOVE_ID,))
move_state = cur.fetchone()[0]
log(f"  → Trạng thái hiện tại: {move_state}")
if move_state == 'posted':
    xmodels.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
        'account.move', 'button_draft', [[MISC_MOVE_ID]])
    log("  → Đã chuyển về Draft")
elif move_state == 'draft':
    log("  → Đã ở Draft, bỏ qua bước reset")

# Lấy dòng TK 1121 trong bút toán MISC
cur.execute("""
    SELECT id, debit, credit FROM account_move_line
    WHERE move_id = %s AND account_id = %s;
""", (MISC_MOVE_ID, acc_1121_id))
lines_1121 = cur.fetchall()
log(f"  → Dòng TK 1121 trong MISC: {lines_1121}")

# Lấy dòng TK 4111 (dòng cân đối)
cur.execute("""
    SELECT id, debit, credit FROM account_move_line
    WHERE move_id = %s AND account_id = %s;
""", (MISC_MOVE_ID, balance_acc_id))
lines_4111 = cur.fetchall()
log(f"  → Dòng TK 4111 trong MISC: {lines_4111}")

# Tính tổng Nợ/Có của dòng 1121 cần xóa
total_1121_debit  = sum(r[1] for r in lines_1121)
total_1121_credit = sum(r[2] for r in lines_1121)

# Xóa dòng TK 1121 VÀ điều chỉnh TK 4111 trong một lệnh write duy nhất (tránh lỗi mất cân bằng)
line_ids_to_delete = [r[0] for r in lines_1121]
if line_ids_to_delete and lines_4111:
    old_credit = float(lines_4111[0][2])
    new_credit = old_credit - float(total_1121_debit) + float(total_1121_credit)
    line_4111_id = lines_4111[0][0]

    # Gộp: xóa dòng 1121 + cập nhật dòng 4111 → 1 lần write
    cmds = [(2, lid, False) for lid in line_ids_to_delete]
    cmds.append((1, line_4111_id, {'credit': new_credit}))
    xmodels.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
        'account.move', 'write',
        [[MISC_MOVE_ID], {'line_ids': cmds}])
    log(f"  ✅ Đã xóa {len(line_ids_to_delete)} dòng TK 1121 và điều chỉnh TK 4111: {old_credit:,.0f} → {new_credit:,.0f}")

# Post lại bút toán MISC
xmodels.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
    'account.move', 'action_post', [[MISC_MOVE_ID]])
log(f"  ✅ Đã post lại bút toán MISC ID={MISC_MOVE_ID}")

# ─── 5. Tạo bút toán cho từng tài khoản ngân hàng ────────────────────────────
log(f"\n🏦 Tạo {len(matched)} bút toán ngân hàng...")
created_moves = []

for item in matched:
    b = item['bank']
    j = item['journal']
    balance = b['balance']
    gl_acc_id = j['gl_acc_id']   # TK 1121 của journal này

    if balance > 0:
        # Dư Nợ: Dr TK1121 / Cr TK4111
        lines = [
            {'account_id': gl_acc_id,     'name': f"Số dư đầu kỳ {b['acc_num']}", 'debit': balance,  'credit': 0.0},
            {'account_id': balance_acc_id, 'name': f"Vốn chủ sở hữu đầu kỳ",      'debit': 0.0,      'credit': balance},
        ]
    else:
        # Dư Có (âm): Cr TK1121 / Dr TK4111
        abs_bal = abs(balance)
        lines = [
            {'account_id': balance_acc_id, 'name': f"Vốn chủ sở hữu đầu kỳ",      'debit': abs_bal, 'credit': 0.0},
            {'account_id': gl_acc_id,     'name': f"Số dư đầu kỳ {b['acc_num']}", 'debit': 0.0,     'credit': abs_bal},
        ]

    move_vals = {
        'journal_id': j['journal_id'],
        'date'       : OPENING_DATE,
        'ref'        : f"Số dư đầu kỳ - {b['acc_num']} - {b['bank_name'][:40]}",
        'move_type'  : 'entry',
        'line_ids'   : [(0, 0, line) for line in lines],
    }

    move_id = xmodels.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
        'account.move', 'create', [move_vals])
    xmodels.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
        'account.move', 'action_post', [[move_id]])
    created_moves.append(move_id)
    sign = "Nợ" if balance >= 0 else "Có"
    log(f"  ✅ {j['code']:6} Move ID={move_id:3} | {b['acc_num']:35} | {sign} {abs(balance):>15,.0f}")

conn.commit()
cur.close(); conn.close()

log(f"""
╔══════════════════════════════════════════════════════════╗
║       HOÀN TẤT NHẬP SỐ DƯ TÀI KHOẢN NGÂN HÀNG          ║
╠══════════════════════════════════════════════════════════╣
║  Bút toán đã tạo    : {len(created_moves):<35} ║
║  Tổng số dư 1121    : {total_balance:>20,.0f}           ║
║  Ngày               : {OPENING_DATE:<35} ║
║  MISC ID={MISC_MOVE_ID} đã xóa dòng 1121 và điều chỉnh TK 4111  ║
╚══════════════════════════════════════════════════════════╝
""")
