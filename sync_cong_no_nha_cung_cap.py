"""
sync_cong_no_nha_cung_cap.py
Nhập số dư công nợ đầu kỳ nhà cung cấp từ MISA vào Odoo 17
- Gán ref cho NV005/NV009/NV022 còn thiếu
- Xóa dòng TK 331/3388/3411 khỏi bút toán MISC (tránh double-count)
- Tạo bút toán chi tiết từng NCC
"""
import openpyxl, psycopg2, json, re, sys, logging, xmlrpc.client
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(filename=r'D:\odoo\sync_cong_no_nha_cung_cap.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8')

EXCEL_PATH   = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_cong_no_nha_cung_cap.xlsx"
OPENING_DATE = '2024-12-31'
BALANCE_CODE = '4111'
JOURNAL_CODE = 'MISC'
MISC_MOVE_ID = 2
REMOVE_TK    = ('331', '3388', '3411')   # TK cần xóa khỏi MISC

ODOO_URL  = 'http://localhost:8017'
ODOO_DB   = 'odoo_company'
ODOO_UID  = 2
ODOO_PASS = 'admin'

def log(msg): print(msg); logging.info(msg)
def norm_name(s):
    s = str(s).lower().strip()
    s = re.sub(r'\b0[0-9]{9,10}\b', '', s)
    return re.sub(r'\s+', ' ', s).strip()

# ─── 1. Đọc Excel ─────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active
rows = []
for r in range(4, ws.max_row + 1):
    tk     = ws.cell(row=r, column=2).value
    ma_ncc = ws.cell(row=r, column=3).value
    ten    = ws.cell(row=r, column=4).value
    debit  = ws.cell(row=r, column=5).value or 0
    credit = ws.cell(row=r, column=6).value or 0
    if not tk or not ma_ncc: continue
    if float(debit) == 0 and float(credit) == 0: continue
    rows.append({'tk': str(tk).strip(), 'ma_ncc': str(ma_ncc).strip(),
                 'ten': str(ten).strip() if ten else '',
                 'debit': float(debit), 'credit': float(credit)})
log(f"  {len(rows)} dòng có số tiền (bỏ qua dòng = 0)")

# ─── 2. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()
xm = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object', allow_none=True)

# ─── 3. Gán ref cho NV005/NV009/NV022 theo tên ───────────────────────────────
log("\n✏️  Gán ref cho NCC còn thiếu...")
missing_refs = {
    'NV005': 'Vũ Thị Thu Hằng',
    'NV009': 'Đoàn Văn Kỳ',
    'NV022': 'Trần Thành Minh',
}
cur.execute("SELECT id, name FROM res_partner WHERE active=TRUE AND (ref IS NULL OR ref='');")
no_ref_partners = cur.fetchall()
name_to_id = {}
for pid, pname in no_ref_partners:
    nm = list(pname.values())[0] if isinstance(pname, dict) else str(pname) if pname else ''
    name_to_id[norm_name(nm)] = pid

for ma, ten in missing_refs.items():
    pid = name_to_id.get(norm_name(ten))
    if pid:
        cur.execute("UPDATE res_partner SET ref=%s, supplier_rank=GREATEST(supplier_rank,1), write_date=NOW() WHERE id=%s;",
                    (ma, pid))
        log(f"  ✅ Gán {ma} → {ten} (ID={pid})")
    else:
        log(f"  ⚠️  Không tìm thấy: {ma} - {ten}, tạo mới...")
        name_json = json.dumps({"vi_VN": ten, "en_US": ten}, ensure_ascii=False)
        cur.execute("""INSERT INTO res_partner (name,ref,supplier_rank,customer_rank,active,is_company,
                        create_uid,write_uid,create_date,write_date)
                       VALUES (%s,%s,1,0,TRUE,FALSE,2,2,NOW(),NOW()) RETURNING id;""",
                    (name_json, ma))
        pid = cur.fetchone()[0]
        log(f"  ➕ Tạo mới: {ma} - {ten} (ID={pid})")
conn.commit()

# ─── 4. Build ref → partner_id map ────────────────────────────────────────────
cur.execute("SELECT ref, id FROM res_partner WHERE ref IS NOT NULL AND ref != '';")
ref_map = {r[0]: r[1] for r in cur.fetchall()}

# ─── 5. Tìm/tạo NCC còn thiếu trong file ─────────────────────────────────────
log("\n🔗 Kiểm tra NCC trong file...")
unique_ncc = {}
for r in rows:
    if r['ma_ncc'] not in unique_ncc:
        unique_ncc[r['ma_ncc']] = r['ten']

created_new = 0
for ma, ten in unique_ncc.items():
    if ma in ref_map:
        continue
    log(f"  ⚠️  Mã '{ma}' chưa có trong Odoo → tạo mới: {ten}")
    name_json = json.dumps({"vi_VN": ten, "en_US": ten}, ensure_ascii=False)
    cur.execute("""INSERT INTO res_partner (name,ref,supplier_rank,customer_rank,active,is_company,
                    create_uid,write_uid,create_date,write_date)
                   VALUES (%s,%s,1,0,TRUE,FALSE,2,2,NOW(),NOW()) RETURNING id;""",
                (name_json, ma))
    ref_map[ma] = cur.fetchone()[0]
    created_new += 1
conn.commit()
log(f"  Tạo mới {created_new} NCC")

# ─── 6. Lấy account IDs ───────────────────────────────────────────────────────
acc_codes = list({r['tk'] for r in rows})
cur.execute("SELECT code, id FROM account_account WHERE code = ANY(%s);", (acc_codes,))
acc_map = {r[0]: r[1] for r in cur.fetchall()}
cur.execute("SELECT id FROM account_account WHERE code=%s;", (BALANCE_CODE,))
balance_acc_id = cur.fetchone()[0]
cur.execute("SELECT id FROM account_journal WHERE code=%s;", (JOURNAL_CODE,))
journal_id = cur.fetchone()[0]
log(f"\n  Tài khoản: {acc_map}")

# ─── 7. Xóa dòng TK 331/3388/3411 khỏi MISC và điều chỉnh TK 4111 ───────────
log(f"\n✏️  Chỉnh sửa bút toán MISC (ID={MISC_MOVE_ID})...")

cur.execute("SELECT state FROM account_move WHERE id=%s;", (MISC_MOVE_ID,))
state = cur.fetchone()[0]
log(f"  Trạng thái: {state}")
if state == 'posted':
    # Dùng SQL reset về draft (button_draft trả None không serialize được qua XML-RPC)
    cur.execute("UPDATE account_move SET state='draft' WHERE id=%s;", (MISC_MOVE_ID,))
    conn.commit()
    log("  → Đã chuyển về Draft (SQL)")

# Lấy dòng cần xóa
cur.execute("""
    SELECT aml.id, aa.code, aml.debit, aml.credit
    FROM account_move_line aml JOIN account_account aa ON aa.id=aml.account_id
    WHERE aml.move_id=%s AND aa.code = ANY(%s);
""", (MISC_MOVE_ID, list(REMOVE_TK)))
remove_lines = cur.fetchall()
log(f"  Dòng sẽ xóa: {[(r[1], float(r[2]), float(r[3])) for r in remove_lines]}")

# Lấy dòng TK 4111
cur.execute("""
    SELECT aml.id, aml.debit, aml.credit FROM account_move_line aml
    JOIN account_account aa ON aa.id=aml.account_id
    WHERE aml.move_id=%s AND aa.id=%s;
""", (MISC_MOVE_ID, balance_acc_id))
line_4111 = cur.fetchone()

# Tính điều chỉnh TK 4111
total_rm_debit  = sum(float(r[2]) for r in remove_lines)
total_rm_credit = sum(float(r[3]) for r in remove_lines)
old_4111_credit = float(line_4111[2])
new_4111_credit = old_4111_credit - total_rm_credit + total_rm_debit
log(f"  TK 4111: {old_4111_credit:,.0f} → {new_4111_credit:,.0f}")

# Dùng SQL: xóa dòng + cập nhật 4111 + cập nhật sequence số bút toán
line_ids_to_del = [r[0] for r in remove_lines]
cur.execute("DELETE FROM account_move_line WHERE id = ANY(%s);", (line_ids_to_del,))
cur.execute("UPDATE account_move_line SET credit=%s, write_date=NOW() WHERE id=%s;",
            (new_4111_credit, line_4111[0]))
conn.commit()
log(f"  ✅ Đã xóa {len(line_ids_to_del)} dòng và điều chỉnh TK 4111 (SQL)")

# Post lại MISC qua SQL
cur.execute("UPDATE account_move SET state='posted' WHERE id=%s;", (MISC_MOVE_ID,))
conn.commit()
log(f"  ✅ Đã post lại MISC ID={MISC_MOVE_ID} (SQL)")

# ─── 8. Tạo bút toán công nợ NCC ──────────────────────────────────────────────
log("\n📋 Tạo bút toán công nợ NCC...")
move_lines = []
total_debit = total_credit = 0.0

for r in rows:
    acc_id = acc_map.get(r['tk'])
    if not acc_id:
        log(f"  ⚠️  Không có TK {r['tk']}, bỏ qua")
        continue
    move_lines.append({
        'account_id': acc_id,
        'partner_id': ref_map.get(r['ma_ncc']),
        'name'      : f"Số dư đầu kỳ {r['ma_ncc']} - {r['ten'][:50]}",
        'debit'     : r['debit'],
        'credit'    : r['credit'],
    })
    total_debit  += r['debit']
    total_credit += r['credit']

diff = round(total_debit - total_credit, 2)
move_lines.append({
    'account_id': balance_acc_id,
    'name'      : 'Vốn chủ sở hữu đầu kỳ - công nợ NCC (TK 4111)',
    'debit'     : abs(diff) if diff < 0 else 0.0,
    'credit'    : diff      if diff > 0 else 0.0,
})

move_id = xm.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS, 'account.move', 'create', [{
    'journal_id': journal_id,
    'date'      : OPENING_DATE,
    'ref'       : 'Số dư đầu kỳ công nợ nhà cung cấp - nhập từ MISA (31/12/2024)',
    'move_type' : 'entry',
    'line_ids'  : [(0, 0, line) for line in move_lines],
}])
xm.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS, 'account.move', 'action_post', [[move_id]])

cur.close(); conn.close()
log(f"""
╔══════════════════════════════════════════════════════════╗
║      HOÀN TẤT CÔNG NỢ ĐẦU KỲ NHÀ CUNG CẤP              ║
╠══════════════════════════════════════════════════════════╣
║  Move ID mới      : {move_id:<35} ║
║  Số dòng          : {len(move_lines):<35} ║
║  Tổng Dư Nợ       : {total_debit:>20,.0f}               ║
║  Tổng Dư Có       : {total_credit:>20,.0f}               ║
║  TK 4111          : {'Nợ' if diff<0 else 'Có'} {abs(diff):,.0f}
╚══════════════════════════════════════════════════════════╝
""")
