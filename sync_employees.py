"""
sync_employees.py
-----------------
Import nhân viên từ MISA → Odoo hr.employee
- Tạo phòng ban (department) tự động nếu chưa có
- Nếu NV cũng là KH/NCC → cập nhật res.partner tương ứng
- Chạy lại an toàn (không tạo trùng)
"""
import openpyxl, xmlrpc.client, sys, re
from datetime import datetime, date
sys.stdout.reconfigure(encoding='utf-8')

# ===================== CẤU HÌNH =====================
EXCEL_FILE = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nhan_vien.xlsx"
SHEET_NAME = 'DANH SÁCH NHÂN VIÊN'
DATA_START = 4

ODOO_URL   = 'http://localhost:8017'
ODOO_DB    = 'odoo_company'
ODOO_USER  = 'doanvanky36k21@gmail.com'
ODOO_PASS  = 'admin'
LOG_FILE   = r"D:\odoo\sync_employees.log"
EXT_MODULE = '__import__'
# ====================================================

def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

def clean(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s in ('None', '#N/A') else s

def parse_date(val):
    if not val: return False
    try:
        if isinstance(val, (datetime, date)):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(s.split(' ')[0], fmt.split(' ')[0]).strftime('%Y-%m-%d')
            except: pass
    except: pass
    return False

# ---- Kết nối ----
common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
if not uid:
    log("Lỗi đăng nhập!"); exit(1)
models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
log("Kết nối Odoo OK")

# ---- Kiểm tra module hr ----
hr_installed = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'ir.module.module', 'search_read',
    [[['name','=','hr'],['state','=','installed']]],
    {'fields':['name'], 'limit':1})
if not hr_installed:
    log("Module HR chưa cài! Đang cài..."); exit(1)

# ---- Cache departments ----
dept_records = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'hr.department', 'search_read', [[]], {'fields':['id','name']})
dept_map = {r['name'].strip(): r['id'] for r in dept_records}
log(f"Có {len(dept_map)} phòng ban hiện tại")

def get_or_create_dept(ten_dv, ma_dv):
    key = ten_dv.strip()
    if key in dept_map:
        return dept_map[key]
    dept_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
        'hr.department', 'create', [{'name': ten_dv}])
    dept_map[key] = dept_id
    log(f"  + Tạo phòng ban: [{ma_dv}] {ten_dv}")
    return dept_id

# ---- Ext IDs đã có (employee) ----
existing_emp = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'ir.model.data', 'search_read',
    [[['module','=',EXT_MODULE],['model','=','hr.employee']]],
    {'fields':['name','res_id']})
emp_ext_map = {r['name']: r['res_id'] for r in existing_emp}

# ---- Ext IDs đã có (partner) ----
existing_partner = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
    'ir.model.data', 'search_read',
    [[['module','=',EXT_MODULE],['model','=','res.partner']]],
    {'fields':['name','res_id']})
partner_ext_map = {r['name']: r['res_id'] for r in existing_partner}

# ---- Đọc Excel ----
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]
employees = []
for r in range(DATA_START, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 11)]
    stt, ma_nv, ten_nv, gioi_tinh, ngay_sinh, chuc_danh, ma_dv, ten_dv, la_kh, la_ncc = row
    ma_nv  = clean(ma_nv)
    ten_nv = clean(ten_nv)
    if not ma_nv or not ten_nv:
        continue
    employees.append({
        'ext_id':     f"misa_nv_{re.sub(r'[^a-zA-Z0-9_]', '_', ma_nv)}",
        'ma_nv':      ma_nv,
        'name':       ten_nv,
        'gender':     'male' if clean(gioi_tinh) in ('Nam','nam') else 'female',
        'birthday':   parse_date(ngay_sinh),
        'job_title':  clean(chuc_danh),
        'ma_dv':      clean(ma_dv),
        'ten_dv':     clean(ten_dv),
        'la_kh':      clean(la_kh) == '✓',
        'la_ncc':     clean(la_ncc) == '✓',
    })

log(f"Đọc được {len(employees)} nhân viên")
log("=" * 55)

created = updated = 0
errors  = []

for emp in employees:
    try:
        # 1. Tạo/lấy department
        dept_id = False
        if emp['ten_dv']:
            dept_id = get_or_create_dept(emp['ten_dv'], emp['ma_dv'])

        # 2. Tạo/cập nhật hr.employee
        emp_vals = {
            'name':      emp['name'],
            'gender':    emp['gender'],
            'job_title': emp['job_title'],
        }
        if dept_id:         emp_vals['department_id'] = dept_id
        if emp['birthday']: emp_vals['birthday']      = emp['birthday']

        if emp['ext_id'] in emp_ext_map:
            emp_id = emp_ext_map[emp['ext_id']]
            models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'hr.employee', 'write', [[emp_id], emp_vals])
            updated += 1
            log(f"  ↻ Cập nhật: [{emp['ma_nv']}] {emp['name']}")
        else:
            emp_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'hr.employee', 'create', [emp_vals])
            models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'ir.model.data', 'create', [{
                    'module': EXT_MODULE, 'name': emp['ext_id'],
                    'model': 'hr.employee', 'res_id': emp_id, 'noupdate': True,
                }])
            created += 1
            log(f"  ✓ Tạo mới: [{emp['ma_nv']}] {emp['name']} | {emp['ten_dv']}")

        # 3. Nếu là KH/NCC → cập nhật res.partner
        partner_ext = f"misa_{re.sub(r'[^a-zA-Z0-9_]', '_', emp['ma_nv'])}"
        if emp['la_kh'] or emp['la_ncc']:
            if partner_ext in partner_ext_map:
                p_id = partner_ext_map[partner_ext]
                p_vals = {}
                if emp['la_kh']:  p_vals['customer_rank'] = 1
                if emp['la_ncc']: p_vals['supplier_rank'] = 1
                models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                    'res.partner', 'write', [[p_id], p_vals])
            else:
                # Tạo partner mới cho NV này
                p_vals = {
                    'name':          emp['name'],
                    'company_type':  'person',
                    'customer_rank': 1 if emp['la_kh'] else 0,
                    'supplier_rank': 1 if emp['la_ncc'] else 0,
                    'comment':       f"MISA NV: {emp['ma_nv']} | {emp['job_title']} | {emp['ten_dv']}",
                }
                p_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                    'res.partner', 'create', [p_vals])
                models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                    'ir.model.data', 'create', [{
                        'module': EXT_MODULE, 'name': partner_ext,
                        'model': 'res.partner', 'res_id': p_id, 'noupdate': True,
                    }])
            log(f"    → Cập nhật partner KH/NCC: {emp['name']}")

    except Exception as e:
        errors.append(f"[{emp['ma_nv']}] {emp['name']}: {e}")
        log(f"  ✗ Lỗi [{emp['ma_nv']}] {emp['name']}: {e}")

log("=" * 55)
log(f"KẾT QUẢ:")
log(f"  Tạo mới nhân viên : {created}")
log(f"  Cập nhật          : {updated}")
log(f"  Lỗi               : {len(errors)}")
if errors:
    for e in errors: log(f"    {e}")
log("ĐỒNG BỘ NHÂN VIÊN HOÀN TẤT")
log("=" * 55)
