"""
sync_products.py
Import vào Odoo 17:
  1. Nhóm vật tư/hàng hóa → product.category
  2. Kho → stock.location (dưới WH/Stock)
  3. Đơn vị tính còn thiếu → uom.uom
  4. Hàng hóa / Dịch vụ → product.template + product.product
"""
import openpyxl, psycopg2, json, re, sys, logging
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    filename=r'D:\odoo\sync_products.log', filemode='w',
    level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s',
    encoding='utf-8'
)

FILE_KHO    = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_kho.xlsx"
FILE_NHOM   = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nhom_vat_tu_hang_hoa_dich_vu.xlsx"
FILE_HH     = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_hang_hoa_dich_vu.xlsx"

def log(msg):
    print(msg); logging.info(msg)

conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()

cur.execute("SELECT id FROM res_company ORDER BY id LIMIT 1;")
company_id = cur.fetchone()[0]

# ─── BƯỚC 1: Nhóm VTHH → product.category ──────────────────────────────���────
log("=" * 60)
log("BƯỚC 1: Import nhóm vật tư/hàng hóa → product.category")

wb = openpyxl.load_workbook(FILE_NHOM, data_only=True)
ws = wb.active
nhom_list = []
for r in range(4, ws.max_row + 1):
    ma   = str(ws.cell(row=r, column=2).value or '').strip()
    ten  = str(ws.cell(row=r, column=3).value or '').strip()
    stat = str(ws.cell(row=r, column=4).value or '').strip()
    if ma and ten and stat != 'Ngừng sử dụng':
        nhom_list.append({'ma': ma, 'ten': ten})

log(f"  Đọc được {len(nhom_list)} nhóm")

# Lấy parent_id = "All" (id=1)
cur.execute("SELECT id FROM product_category WHERE name='All' OR name::text ILIKE '%All%' ORDER BY id LIMIT 1;")
parent_categ_id = cur.fetchone()[0]
log(f"  Parent category: id={parent_categ_id}")

categ_map = {}  # ma_nhom → categ_id
# Load existing — name là JSONB nên psycopg2 trả về dict
cur.execute("SELECT id, name FROM product_category;")
for r in cur.fetchall():
    name_data = r[1]  # Python dict hoặc str
    vals = list(name_data.values()) if isinstance(name_data, dict) else [str(name_data)]
    for nhom in nhom_list:
        if nhom['ten'] in vals or nhom['ma'] in vals:
            categ_map[nhom['ma']] = r[0]

created_categ = 0
for nhom in nhom_list:
    if nhom['ma'] in categ_map:
        continue
    name_json = json.dumps({"vi_VN": nhom['ten'], "en_US": nhom['ten']}, ensure_ascii=False)
    cur.execute("""
        INSERT INTO product_category (name, parent_id, create_uid, write_uid, create_date, write_date)
        VALUES (%s, %s, 2, 2, NOW(), NOW()) RETURNING id;
    """, (name_json, parent_categ_id))
    cid = cur.fetchone()[0]
    categ_map[nhom['ma']] = cid
    created_categ += 1
    log(f"  ✅ Tạo category: {nhom['ma']} - {nhom['ten']} (id={cid})")

# Cập nhật complete_name
for nhom in nhom_list:
    if nhom['ma'] in categ_map:
        cid = categ_map[nhom['ma']]
        cur.execute("""
            UPDATE product_category SET parent_path = %s WHERE id = %s;
        """, (f"{parent_categ_id}/{cid}/", cid))

log(f"  → Tạo mới {created_categ}, tổng map {len(categ_map)} nhóm")

# ─── BƯỚC 2: Kho → stock.location ───────────────────────────────────────────
log("\n" + "=" * 60)
log("BƯỚC 2: Import kho → stock.location (dưới WH/Stock)")

wb2 = openpyxl.load_workbook(FILE_KHO, data_only=True)
ws2 = wb2.active
kho_list = []
for r in range(4, ws2.max_row + 1):
    ma   = str(ws2.cell(row=r, column=2).value or '').strip()
    ten  = str(ws2.cell(row=r, column=3).value or '').strip()
    addr = str(ws2.cell(row=r, column=4).value or '').strip()
    stat = str(ws2.cell(row=r, column=5).value or '').strip()
    if ma and ten and stat != 'Ngừng sử dụng':
        kho_list.append({'ma': ma, 'ten': ten, 'addr': addr})

log(f"  Đọc được {len(kho_list)} kho")

# Parent location = WH/Stock (id=8)
cur.execute("SELECT id FROM stock_location WHERE complete_name='WH/Stock' LIMIT 1;")
row = cur.fetchone()
parent_loc_id = row[0] if row else 8
log(f"  Parent location: id={parent_loc_id} (WH/Stock)")

loc_map = {}  # ma_kho → location_id
# Load existing locations
cur.execute("SELECT id, name::text FROM stock_location WHERE usage='internal' AND location_id=%s;", (parent_loc_id,))
for r in cur.fetchall():
    for kho in kho_list:
        if kho['ma'] in r[1] or kho['ten'] in r[1]:
            loc_map[kho['ma']] = r[0]

created_loc = 0
for kho in kho_list:
    if kho['ma'] in loc_map:
        log(f"  ⏩ Đã có: {kho['ma']} - {kho['ten']}")
        continue
    name_json = json.dumps({"vi_VN": kho['ten'], "en_US": kho['ten']}, ensure_ascii=False)
    cur.execute("""
        INSERT INTO stock_location
            (name, complete_name, location_id, usage, active, company_id,
             create_uid, write_uid, create_date, write_date)
        VALUES (%s, %s, %s, 'internal', TRUE, %s, 2, 2, NOW(), NOW())
        RETURNING id;
    """, (name_json, name_json, parent_loc_id, company_id))
    lid = cur.fetchone()[0]
    # Update complete_name properly
    cur.execute("SELECT complete_name FROM stock_location WHERE id=%s;", (parent_loc_id,))
    parent_cn = cur.fetchone()[0] if cur.rowcount else 'WH/Stock'
    cur.execute("UPDATE stock_location SET complete_name=%s WHERE id=%s;",
                (f"{parent_cn}/{kho['ten']}", lid))
    loc_map[kho['ma']] = lid
    created_loc += 1
    log(f"  ✅ Tạo location: {kho['ma']} - {kho['ten']} (id={lid})")

log(f"  → Tạo mới {created_loc}, tổng map {len(loc_map)} kho")

# ─── BƯỚC 3: Đơn vị tính ─────────────────────────────────────────────────────
log("\n" + "=" * 60)
log("BƯỚC 3: Tạo đơn vị tính còn thiếu")

# Lấy UOM đơn vị tính Việt Nam cần tạo
NEED_UOM = ['Bao', 'Bánh', 'Bộ', 'Cái', 'Cây', 'Chai', 'Gói', 'Hộp',
            'Lon', 'Ram', 'Tuýp', 'Túi', 'Túi (100 cái)', 'Quả', 'quả',
            'Lít', 'Hộp', 'kg', 'm3']
# Normalize map MISA name → Odoo uom name
UOM_NORMALIZE = {
    'Lít': 'Lít', 'lít': 'Lít',
    'kg': 'kg', 'Kg': 'kg',
    'm3': 'm³', 'M3': 'm³',
    'quả': 'Quả',
}

# Load UOM hiện có
cur.execute("SELECT id, name::text FROM uom_uom WHERE active=TRUE;")
uom_map = {}  # name (lower) → id
for r in cur.fetchall():
    nd = r[1]
    # Extract name from JSON
    try:
        d = json.loads(nd)
        for v in d.values():
            uom_map[v.lower()] = r[0]
    except Exception:
        uom_map[nd.lower()] = r[0]

log(f"  UOM hiện có: {len(uom_map)}")

# UOM category "Unit" id
cur.execute("SELECT id FROM uom_category WHERE name::text ILIKE '%Unit%' OR name::text ILIKE '%Đơn vị%' LIMIT 1;")
uom_categ_unit = cur.fetchone()[0]  # 1

created_uom = 0
uom_vi_map = {}  # misa_name → uom_id

# Đọc tất cả dvt từ file hàng hóa
wb_hh = openpyxl.load_workbook(FILE_HH, data_only=True)
ws_hh = wb_hh.active
all_dvt = set()
for r in range(4, ws_hh.max_row + 1):
    v = str(ws_hh.cell(row=r, column=6).value or '').strip()
    if v and v not in ('None', ''):
        all_dvt.add(v)

log(f"  Đơn vị tính trong file HH: {sorted(all_dvt)}")

for dvt in sorted(all_dvt):
    # Normalize
    normalized = UOM_NORMALIZE.get(dvt, dvt)
    key = normalized.lower()
    if key in uom_map:
        uom_vi_map[dvt] = uom_map[key]
        continue
    # Tạo mới
    name_json = json.dumps({"vi_VN": normalized, "en_US": normalized}, ensure_ascii=False)
    cur.execute("""
        INSERT INTO uom_uom (name, category_id, uom_type, factor, rounding, active,
                             create_uid, write_uid, create_date, write_date)
        VALUES (%s, %s, 'reference', 1.0, 0.01, TRUE, 2, 2, NOW(), NOW())
        RETURNING id;
    """, (name_json, uom_categ_unit))
    uid = cur.fetchone()[0]
    uom_vi_map[dvt] = uid
    uom_map[key] = uid
    created_uom += 1
    log(f"  ✅ Tạo UOM: {dvt} → {normalized} (id={uid})")

# UOM mặc định cho sản phẩm không có đơn vị (Dịch vụ)
DEFAULT_UOM_ID = uom_map.get('units') or uom_map.get('đơn vị') or 1
log(f"  → Tạo mới {created_uom} UOM, default UOM id={DEFAULT_UOM_ID}")

# ─── BƯỚC 4: Import sản phẩm ─────────────────────────────────────────────────
log("\n" + "=" * 60)
log("BƯỚC 4: Import sản phẩm/dịch vụ → product.template")

# Load tài khoản để map
cur.execute("SELECT id, code FROM account_account WHERE company_id=%s;", (company_id,))
acct_map = {r[1]: r[0] for r in cur.fetchall()}

# Đọc sản phẩm hiện có (theo default_code)
cur.execute("SELECT id, default_code FROM product_template WHERE company_id=%s OR company_id IS NULL;", (company_id,))
existing_products = {r[1]: r[0] for r in cur.fetchall() if r[1]}
log(f"  Sản phẩm đã có trong Odoo: {len(existing_products)}")

# Category mặc định
categ_default = parent_categ_id  # All

created_prod = 0
updated_prod = 0
skipped_prod = 0

for r in range(4, ws_hh.max_row + 1):
    ma   = str(ws_hh.cell(row=r, column=2).value or '').strip()
    ten  = str(ws_hh.cell(row=r, column=3).value or '').strip()
    tc   = str(ws_hh.cell(row=r, column=4).value or '').strip()   # Tính chất
    nhom = str(ws_hh.cell(row=r, column=5).value or '').strip()   # Nhóm VTHH
    dvt  = str(ws_hh.cell(row=r, column=6).value or '').strip()   # ĐVT
    # tk_kho     = ws_hh.cell(row=r, column=11).value
    tk_dt      = str(ws_hh.cell(row=r, column=12).value or '').strip()  # TK Doanh thu
    tk_cp      = str(ws_hh.cell(row=r, column=15).value or '').strip()  # TK Chi phí

    if not ma or not ten:
        continue

    if ma in existing_products:
        skipped_prod += 1
        continue

    # Loại sản phẩm
    if tc == 'Dịch vụ':
        detailed_type = 'service'
        prod_type     = 'service'
    else:  # Hàng hóa, CCDC, NVL, TP
        detailed_type = 'product'   # storable
        prod_type     = 'product'

    # Category
    categ_id = categ_map.get(nhom, categ_default)

    # UOM
    uom_id = uom_vi_map.get(dvt, DEFAULT_UOM_ID) if dvt else DEFAULT_UOM_ID

    name_json = json.dumps({"vi_VN": ten, "en_US": ten}, ensure_ascii=False)

    try:
        cur.execute("SAVEPOINT sp1;")
        cur.execute("""
            INSERT INTO product_template
                (name, default_code, detailed_type, type, categ_id,
                 uom_id, uom_po_id, active, sale_ok, purchase_ok,
                 sale_line_warn, purchase_line_warn, tracking,
                 company_id, create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, %s, %s, %s,
                    %s, %s, TRUE, TRUE, TRUE,
                    'no-message', 'no-message', 'none',
                    %s, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (name_json, ma, detailed_type, prod_type, categ_id,
              uom_id, uom_id, company_id))
        tmpl_id = cur.fetchone()[0]

        # Tạo product.product (variant)
        cur.execute("""
            INSERT INTO product_product
                (product_tmpl_id, active, create_uid, write_uid, create_date, write_date)
            VALUES (%s, TRUE, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (tmpl_id,))

        cur.execute("RELEASE SAVEPOINT sp1;")
        created_prod += 1
        if created_prod % 100 == 0:
            log(f"  ... đã tạo {created_prod} sản phẩm")

    except Exception as e:
        cur.execute("ROLLBACK TO SAVEPOINT sp1;")
        log(f"  ❌ Lỗi SP {ma}: {e}")
        continue

conn.commit()
log(f"\n✅ HOÀN TẤT!")
log(f"   Category tạo mới : {created_categ}")
log(f"   Kho tạo mới      : {created_loc}")
log(f"   UOM tạo mới      : {created_uom}")
log(f"   Sản phẩm tạo mới : {created_prod}")
log(f"   Sản phẩm đã có   : {skipped_prod}")
cur.close(); conn.close()
