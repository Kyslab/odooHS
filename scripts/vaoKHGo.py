# -*- coding: utf-8 -*-
"""
vaoKHGo.py  —  Đồng bộ Google Sheet → Excel
================================================
Trigger   : nhập lệnh "vaoKHGo" (hoặc gọi hàm run() từ bot khác)
Nguồn     : https://docs.google.com/spreadsheets/d/1nvlMgrv75r6W9aGaMZCTUt4pbBmCAwAk/
Sheet     : sheet có tên = ngày hôm nay (vd: 22/04 → "2204", "22.04", "22/04", "22/4"...)
Sao chép  : dòng 6 → 130
Đích      : D:\\Dropbox\\1. Tien_duong_tu_13.08.2021\\kh nhap GGS4.xlsm  |  Sheet2  |  từ ô B3
Lưu ý     : KHÔNG tự lưu — file Excel sẽ mở sẵn để người dùng kiểm tra rồi tự lưu
"""

import sys
import io
import os
import requests
import openpyxl
from datetime import datetime

# ─── Cấu hình ───────────────────────────────────────────────────────────────
SPREADSHEET_ID   = "1nvlMgrv75r6W9aGaMZCTUt4pbBmCAwAk"
TARGET_FILE      = r"D:\Dropbox\1. Tien_duong_tu_13.08.2021\kh nhap GGS4.xlsm"
TARGET_SHEET     = "Sheet2"
TARGET_START_ROW = 3    # dán từ hàng 3 (ô B3)
TARGET_START_COL = 2    # cột B = 2
SOURCE_ROW_FROM  = 6    # lấy từ dòng 6
SOURCE_ROW_TO    = 130  # đến dòng 130  (tổng = 125 dòng)

EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
)

# ─── Các định dạng tên sheet có thể có theo ngày ────────────────────────────
def get_candidate_sheet_names(dt: datetime) -> list:
    d, m = dt.day, dt.month
    d2 = f"{d:02d}"
    m2 = f"{m:02d}"
    d1 = str(d)
    m1 = str(m)
    return [
        d2 + m2,           # 2204   ← phổ biến nhất
        f"{d2}.{m2}",      # 22.04
        f"{d2}/{m2}",      # 22/04
        f"{d2}.{m1}",      # 22.4
        f"{d2}/{m1}",      # 22/4
        f"{d1}.{m2}",      # 22.04  (không pad)
        f"{d1}/{m2}",      # 22/04
        f"{d1}.{m1}",      # 22.4
        f"{d1}/{m1}",      # 22/4
        d2,                # 22     (chỉ ngày)
        d1,                # 22
    ]

# ─── Tải spreadsheet từ Google ───────────────────────────────────────────────
def download_workbook(url: str, retries: int = 3) -> openpyxl.Workbook:
    print("  📥  Đang tải Google Sheet...", flush=True)
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, timeout=90, allow_redirects=True, stream=False)
            resp.raise_for_status()
            content_type = resp.headers.get("Content-Type", "")
            if "spreadsheetml" not in content_type and "officedocument" not in content_type:
                raise RuntimeError(
                    "Google Sheet không công khai hoặc URL sai.\n"
                    f"  → Vào link và chọn 'Chia sẻ' → 'Mọi người có đường liên kết':\n"
                    f"  → https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/"
                )
            data = resp.content
            print(f"  ✅  Tải xong ({len(data)//1024} KB)", flush=True)
            return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except RuntimeError:
            raise
        except Exception as e:
            last_err = e
            if attempt < retries:
                print(f"  ⚠️   Lần {attempt} lỗi ({e}), thử lại...", flush=True)
    raise RuntimeError(f"Không thể tải Google Sheet sau {retries} lần:\n  {last_err}")

# ─── Tìm sheet hôm nay ───────────────────────────────────────────────────────
def find_today_sheet(wb: openpyxl.Workbook, dt: datetime):
    candidates = get_candidate_sheet_names(dt)
    # So sánh chính xác trước
    name_map = {s.strip(): s for s in wb.sheetnames}
    for c in candidates:
        if c in name_map:
            return name_map[c]
    # So sánh không phân biệt hoa-thường
    for c in candidates:
        for s in wb.sheetnames:
            if s.strip().lower() == c.lower():
                return s
    return None

# ─── Đọc dữ liệu từ sheet nguồn ──────────────────────────────────────────────
def read_source_rows(wb: openpyxl.Workbook, sheet_name: str,
                     row_from: int, row_to: int) -> list:
    ws = wb[sheet_name]
    rows_data = []
    for row in ws.iter_rows(min_row=row_from, max_row=row_to, values_only=True):
        rows_data.append(list(row))
    return rows_data

# ─── Ghi vào Excel qua COM — KHÔNG lưu, file mở sẵn cho người dùng ──────────
def write_to_excel_com(rows_data: list, target_file: str, sheet_name: str,
                       start_row: int, start_col: int) -> int:
    """
    Dùng win32com để mở file Excel thật, dán dữ liệu, rồi để file mở (không lưu).
    Nếu Excel đang mở sẵn file đó → dùng luôn instance đó.
    """
    import win32com.client
    import pythoncom

    if not os.path.exists(target_file):
        raise FileNotFoundError(f"Không tìm thấy file:\n  {target_file}")

    abs_path = os.path.abspath(target_file)
    print(f"  📂  Mở Excel: {os.path.basename(abs_path)}", flush=True)

    pythoncom.CoInitialize()
    try:
        # Thử dùng Excel đang mở — nếu chưa có thì tạo mới
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")

        xl.Visible = True
        xl.DisplayAlerts = False

        # Tìm workbook đã mở, hoặc mở mới
        wb_com = None
        for wb_open in xl.Workbooks:
            if os.path.abspath(wb_open.FullName).lower() == abs_path.lower():
                wb_com = wb_open
                break

        if wb_com is None:
            wb_com = xl.Workbooks.Open(abs_path)

        # Tìm sheet đích
        ws_com = None
        for sh in wb_com.Sheets:
            if sh.Name == sheet_name:
                ws_com = sh
                break
        if ws_com is None:
            raise ValueError(
                f"Sheet '{sheet_name}' không tồn tại trong file.\n"
                f"  Có: {[s.Name for s in wb_com.Sheets]}"
            )

        ws_com.Activate()

        # ── Xóa vùng dán cũ (từ B{start_row} xuống) ───────────────────────
        num_rows    = len(rows_data)
        num_cols    = max((len(r) for r in rows_data), default=0)
        clear_range = ws_com.Range(
            ws_com.Cells(start_row, start_col),
            ws_com.Cells(start_row + num_rows + 4, start_col + num_cols + 1),
        )
        clear_range.ClearContents()

        # ── Ghi từng ô ────────────────────────────────────────────────────
        # Cách nhanh: gán cả mảng 2 chiều một lần (COM array assignment)
        # Chuẩn bị mảng — COM cần tuple of tuples
        com_array = []
        for row_vals in rows_data:
            # Đảm bảo đủ cột
            padded = list(row_vals) + [None] * (num_cols - len(row_vals))
            # Chuyển None → Empty, giữ số và chuỗi
            com_array.append(tuple(
                "" if v is None else v for v in padded
            ))

        target_range = ws_com.Range(
            ws_com.Cells(start_row, start_col),
            ws_com.Cells(start_row + num_rows - 1, start_col + num_cols - 1),
        )
        target_range.Value = com_array

        # ── Kích hoạt ô B{start_row} sau khi dán ──────────────────────────
        ws_com.Cells(start_row, start_col).Select()

        xl.DisplayAlerts = True
        print(
            f"  📋  Đã dán {num_rows} dòng × {num_cols} cột vào "
            f"'{sheet_name}'!{_col_letter(start_col)}{start_row}",
            flush=True,
        )
        print("  ⚠️   File CHƯA được lưu — vui lòng kiểm tra rồi Ctrl+S để lưu.", flush=True)
        return num_rows

    finally:
        pythoncom.CoUninitialize()


def _col_letter(col: int) -> str:
    """Chuyển số cột sang chữ cái (1→A, 2→B, ...)"""
    result = ""
    while col:
        col, r = divmod(col - 1, 26)
        result = chr(65 + r) + result
    return result


# ─── Hàm chính ───────────────────────────────────────────────────────────────
def run(date_override: datetime = None) -> bool:
    dt = date_override or datetime.now()
    date_str = dt.strftime("%d/%m/%Y")
    print(f"\n{'='*55}")
    print(f"  vaoKHGo — đồng bộ Google Sheet → Excel")
    print(f"  Ngày xử lý : {date_str}")
    print(f"{'='*55}")

    try:
        # 1. Tải workbook từ Google
        wb = download_workbook(EXPORT_URL)

        # 2. Tìm sheet hôm nay
        sheet_name = find_today_sheet(wb, dt)
        if not sheet_name:
            candidates = get_candidate_sheet_names(dt)
            available  = ", ".join(wb.sheetnames[-10:])
            raise RuntimeError(
                f"Không tìm thấy sheet cho ngày {date_str}.\n"
                f"  Đã thử: {candidates}\n"
                f"  10 sheet cuối: {available}"
            )

        print(f"  📋  Sheet tìm thấy  : '{sheet_name}'", flush=True)
        print(f"  📌  Lấy dòng {SOURCE_ROW_FROM} → {SOURCE_ROW_TO}", flush=True)

        # 3. Đọc dữ liệu nguồn
        rows_data = read_source_rows(wb, sheet_name, SOURCE_ROW_FROM, SOURCE_ROW_TO)
        wb.close()

        # Cắt bỏ các dòng cuối toàn rỗng
        while rows_data and all(v is None or v == "" for v in rows_data[-1]):
            rows_data.pop()

        print(f"  📊  {len(rows_data)} dòng có dữ liệu", flush=True)

        # 4. Ghi vào Excel (qua COM — không lưu)
        write_to_excel_com(
            rows_data,
            TARGET_FILE,
            TARGET_SHEET,
            TARGET_START_ROW,
            TARGET_START_COL,
        )

        print(f"\n  ✅  HOÀN TẤT! Dữ liệu đã dán vào Sheet2 từ ô B{TARGET_START_ROW}.")
        print(f"      Excel đang mở — kiểm tra xong nhấn Ctrl+S để lưu.")
        print(f"{'='*55}\n")
        return True

    except Exception as e:
        print(f"\n  ❌  LỖI: {e}")
        print(f"{'='*55}\n")
        return False


# ─── Entry point — lắng nghe lệnh "vaoKHGo" ─────────────────────────────────
def main():
    if len(sys.argv) > 1:
        trigger = sys.argv[1].strip()
        if trigger == "vaoKHGo":
            sys.exit(0 if run() else 1)
        else:
            print(f"Lệnh không hợp lệ: '{trigger}'. Dùng: vaoKHGo")
            sys.exit(1)

    # Chế độ tương tác
    print("Bot sẵn sàng. Nhập lệnh (Ctrl+C để thoát):")
    while True:
        try:
            cmd = input(">> ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\nĐã thoát.")
            break

        if cmd == "vaoKHGo":
            run()
        elif cmd in ("exit", "quit", "thoat"):
            print("Đã thoát.")
            break
        elif cmd == "":
            continue
        else:
            print(f"Không nhận ra '{cmd}'. Lệnh hợp lệ: vaoKHGo | exit")


if __name__ == "__main__":
    main()
