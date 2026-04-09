"""
sync_ma_khach_hang.py
Đọc Danh_sach_khach_hang.xlsx → gán Mã KH (ref) cho tất cả partner trong Odoo
Chiến lược khớp (theo thứ tự ưu tiên):
  1. Ref đã khớp → bỏ qua
  2. Khớp số điện thoại
  3. Khớp tên (chuẩn hóa)
  4. Chưa tìm thấy → tạo mới
"""
import openpyxl, psycopg2, re, json, sys, logging
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(filename=r'D:\odoo\sync_ma_khach_hang.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8')

EXCEL_PATH = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khach_hang.xlsx"

def log(msg): print(msg); logging.info(msg)

def norm_phone(p):
    """Chuẩn hóa SĐT: chỉ giữ chữ số, bỏ +84"""
    if not p: return ''
    s = re.sub(r'[^0-9]', '', str(p))
    if s.startswith('84') and len(s) >= 11: s = '0' + s[2:]
    return s

def norm_name(s):
    """Chuẩn hóa tên: lower, bỏ SĐT nhúng, bỏ ký tự đặc biệt thừa"""
    s = str(s).lower().strip()
    s = re.sub(r'\b0[0-9]{9,10}\b', '', s)   # bỏ SĐT 10 số
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ─── 1. Đọc Excel ─────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

file_customers = []   # {'ma_kh', 'ten_kh', 'dia_chi', 'mst', 'phone', 'mobile', 'is_supplier'}
for r in range(4, ws.max_row + 1):
    ma_kh  = ws.cell(row=r, column=2).value
    ten_kh = ws.cell(row=r, column=3).value
    dia_chi= ws.cell(row=r, column=4).value
    mst    = ws.cell(row=r, column=7).value
    phone  = ws.cell(row=r, column=8).value
    mobile = ws.cell(row=r, column=10).value
    is_sup = ws.cell(row=r, column=11).value

    if not ma_kh or not ten_kh: continue
    ma_kh  = str(ma_kh).strip()
    ten_kh = str(ten_kh).strip()
    if ten_kh in ('#N/A', 'Tổng', '') or ma_kh in ('0', '#N/A', 'Tổng'):
        continue
    # Bỏ qua mã không hợp lệ (không theo pattern XXnnnnn)
    if not re.match(r'^[A-Za-z]+[0-9]+$', ma_kh):
        continue

    file_customers.append({
        'ma_kh'      : ma_kh,
        'ten_kh'     : ten_kh,
        'dia_chi'    : str(dia_chi).strip() if dia_chi else '',
        'mst'        : str(mst).strip() if mst else '',
        'phone'      : norm_phone(phone),
        'mobile'     : norm_phone(mobile),
        'is_supplier': bool(is_sup and str(is_sup).strip() in ('✓','x','X','1','True','TRUE')),
    })

log(f"  Đọc được {len(file_customers)} khách hàng hợp lệ từ file")

# ─── 2. Kết nối DB, load tất cả partner ──────────────────────────────────────
log("\n🔗 Kết nối DB...")
conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()

cur.execute("""
    SELECT id, name, ref, phone, mobile, vat
    FROM res_partner WHERE active=TRUE;
""")
odoo_partners = cur.fetchall()
log(f"  Tổng partner trong Odoo: {len(odoo_partners)}")

# Build lookup maps
ref_map    = {}   # ref → partner_id (đã có mã)
phone_map  = {}   # norm_phone → partner_id
name_map   = {}   # norm_name → partner_id

for pid, pname, pref, pphone, pmobile, pvat in odoo_partners:
    nm = list(pname.values())[0] if isinstance(pname, dict) else str(pname) if pname else ''

    if pref:
        ref_map[pref.strip()] = pid

    for ph in [pphone, pmobile]:
        np = norm_phone(ph)
        if np and len(np) >= 9:
            phone_map.setdefault(np, pid)

    nn = norm_name(nm)
    if nn:
        name_map.setdefault(nn, pid)

log(f"  Phone map: {len(phone_map)} entries | Name map: {len(name_map)} entries")

# ─── 3. Khớp và cập nhật ─────────────────────────────────────────────────────
log("\n🔗 Khớp Mã KH...")
matched_ref   = 0   # đã có ref khớp
matched_phone = 0   # khớp qua SĐT
matched_name  = 0   # khớp qua tên
created_new   = 0   # tạo mới
not_matched   = []

updates = []  # (partner_id, ma_kh, dia_chi, mst, phone, mobile, is_supplier)

for c in file_customers:
    ma_kh = c['ma_kh']

    # 1. Đã có ref trùng → chỉ cập nhật thông tin còn thiếu
    if ma_kh in ref_map:
        matched_ref += 1
        updates.append((ref_map[ma_kh], ma_kh, c['dia_chi'], c['mst'],
                        c['phone'], c['mobile'], c['is_supplier']))
        continue

    # 2. Khớp SĐT
    found_id = None
    for ph in [c['phone'], c['mobile']]:
        if ph and ph in phone_map:
            found_id = phone_map[ph]
            break

    if found_id:
        matched_phone += 1
        updates.append((found_id, ma_kh, c['dia_chi'], c['mst'],
                        c['phone'], c['mobile'], c['is_supplier']))
        ref_map[ma_kh] = found_id
        continue

    # 3. Khớp tên (chuẩn hóa)
    nn = norm_name(c['ten_kh'])
    if nn in name_map:
        found_id = name_map[nn]
        matched_name += 1
        updates.append((found_id, ma_kh, c['dia_chi'], c['mst'],
                        c['phone'], c['mobile'], c['is_supplier']))
        ref_map[ma_kh] = found_id
        continue

    # 4. Không tìm thấy → tạo mới
    not_matched.append(c)

log(f"  Đã có ref sẵn   : {matched_ref}")
log(f"  Khớp qua SĐT    : {matched_phone}")
log(f"  Khớp qua tên    : {matched_name}")
log(f"  Cần tạo mới     : {len(not_matched)}")

# ─── 4. Cập nhật DB ───────────────────────────────────────────────────────────
log("\n✏️  Cập nhật partner trong Odoo...")
updated = 0
for (pid, ma_kh, dia_chi, mst, phone, mobile, is_supplier) in updates:
    supplier_rank = 1 if is_supplier else 0
    # Chỉ set ref nếu chưa có, cập nhật các field còn lại nếu còn trống
    cur.execute("""
        UPDATE res_partner SET
            ref          = COALESCE(NULLIF(ref,''), %s),
            street       = CASE WHEN (street IS NULL OR street='') AND %s != '' THEN %s ELSE street END,
            vat          = CASE WHEN (vat  IS NULL OR vat ='') AND %s != '' THEN %s ELSE vat  END,
            phone        = CASE WHEN (phone  IS NULL OR phone ='') AND %s != '' THEN %s ELSE phone  END,
            mobile       = CASE WHEN (mobile IS NULL OR mobile='') AND %s != '' THEN %s ELSE mobile END,
            supplier_rank= GREATEST(supplier_rank, %s),
            write_date   = NOW(), write_uid = 2
        WHERE id = %s;
    """, (ma_kh,
          dia_chi, dia_chi,
          mst, mst,
          phone, phone,
          mobile, mobile,
          supplier_rank,
          pid))
    updated += 1

log(f"  ✅ Đã cập nhật {updated} partner")

# ─── 5. Tạo mới những khách hàng chưa có ─────────────────────────────────────
if not_matched:
    log(f"\n➕ Tạo mới {len(not_matched)} partner...")
    for c in not_matched:
        name_json = json.dumps({"vi_VN": c['ten_kh'], "en_US": c['ten_kh']}, ensure_ascii=False)
        cur.execute("""
            INSERT INTO res_partner
                (name, ref, street, vat, phone, mobile,
                 customer_rank, supplier_rank, active, is_company,
                 create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, NULLIF(%s,''), NULLIF(%s,''), NULLIF(%s,''), NULLIF(%s,''),
                    1, %s, TRUE, FALSE, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (name_json, c['ma_kh'], c['dia_chi'], c['mst'],
              c['phone'] or None, c['mobile'] or None,
              1 if c['is_supplier'] else 0))
        new_id = cur.fetchone()[0]
        ref_map[c['ma_kh']] = new_id
        created_new += 1
        log(f"  ➕ Tạo mới: {c['ma_kh']} - {c['ten_kh']}")

conn.commit()
cur.close(); conn.close()

log(f"""
╔══════════════════════════════════════════════════════════╗
║          HOÀN TẤT GÁN MÃ KHÁCH HÀNG                     ║
╠══════════════════════════════════════════════════════════╣
║  Tổng KH trong file   : {len(file_customers):<33} ║
║  Đã có ref sẵn        : {matched_ref:<33} ║
║  Khớp qua SĐT         : {matched_phone:<33} ║
║  Khớp qua tên         : {matched_name:<33} ║
║  Tạo mới              : {created_new:<33} ║
╚══════════════════════════════════════════════════════════╝
""")
