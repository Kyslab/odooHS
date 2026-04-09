"""
sync_analytic.py
----------------
Import danh sách công trình từ MISA → Odoo Analytic Accounts
"""
import openpyxl, psycopg2, sys, re, json
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_cong_trinh.xlsx"
SHEET_NAME = 'DANH SÁCH CÔNG TRÌNH'
DATA_START = 4
LOG_FILE   = r"D:\odoo\sync_analytic.log"

def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

def clean(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s in ('None', '#N/A') else s

def parse_date(val):
    if not val or str(val).strip() in ('', 'None'): return None
    try:
        if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
        for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d'):
            try: return datetime.strptime(str(val).split(' ')[0], fmt.split(' ')[0]).strftime('%Y-%m-%d')
            except: pass
    except: pass
    return None

# ---- Kết nối DB ----
conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = True
cur = conn.cursor()
log("Kết nối DB OK")

# ---- Lấy hoặc tạo Analytic Plan ----
cur.execute("SELECT id, name FROM account_analytic_plan LIMIT 5;")
plans = cur.fetchall()
log(f"Plans hiện có: {plans}")

if plans:
    plan_id = plans[0][0]
    log(f"Dùng plan id={plan_id}: {plans[0][1]}")
else:
    cur.execute("""
        INSERT INTO account_analytic_plan (name, color, complete_name)
        VALUES ('Công trình / Dự án', 1, 'Công trình / Dự án')
        RETURNING id;
    """)
    plan_id = cur.fetchone()[0]
    log(f"Tạo plan mới id={plan_id}")

# ---- Lấy company_id ----
cur.execute("SELECT id FROM res_company LIMIT 1;")
company_id = cur.fetchone()[0]

# ---- Đọc Excel ----
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]
projects = []
for r in range(DATA_START, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    stt, ma_ct, ten_ct, loai, tinh_trang, ngay_bd, ngay_kt, du_toan, chu_dt, chi_nhanh, trang_thai = row
    ma_ct  = clean(ma_ct)
    ten_ct = clean(ten_ct)
    if not ma_ct or not ten_ct: continue
    ext_key = f"misa_ct_{re.sub(r'[^a-zA-Z0-9_]', '_', ma_ct)}"
    projects.append({
        'ext_key':    ext_key,
        'ma_ct':      ma_ct,
        'name':       ten_ct,
        'date_start': parse_date(ngay_bd),
        'date_end':   parse_date(ngay_kt),
        'budget':     float(du_toan) if du_toan and str(du_toan) not in ('0','None','') else 0,
    })

log(f"Đọc được {len(projects)} công trình")
log("=" * 55)

created = updated = 0
errors  = []

for proj in projects:
    try:
        # Kiểm tra đã tồn tại chưa (theo code)
        cur.execute("SELECT id FROM account_analytic_account WHERE code=%s AND company_id=%s;",
                    (proj['ma_ct'], company_id))
        existing = cur.fetchone()

        # name phải lưu dạng JSON đa ngôn ngữ
        name_json = json.dumps({"vi_VN": proj['name'], "en_US": proj['name']}, ensure_ascii=False)

        if existing:
            acc_id = existing[0]
            cur.execute("""
                UPDATE account_analytic_account
                SET name=%s, plan_id=%s, root_plan_id=%s
                WHERE id=%s;
            """, (name_json, plan_id, plan_id, acc_id))
            updated += 1
            log(f"  ↻ Cập nhật: [{proj['ma_ct']}] {proj['name']}")
        else:
            cur.execute("""
                INSERT INTO account_analytic_account
                    (name, code, plan_id, root_plan_id, company_id, active)
                VALUES (%s, %s, %s, %s, %s, TRUE)
                RETURNING id;
            """, (name_json, proj['ma_ct'], plan_id, plan_id, company_id))
            acc_id = cur.fetchone()[0]
            created += 1
            budget_str = f" | Dự toán: {proj['budget']:,.0f} đ" if proj['budget'] > 0 else ""
            log(f"  ✓ Tạo mới: [{proj['ma_ct']}] {proj['name']}{budget_str}")

    except Exception as e:
        conn.rollback()
        errors.append(f"[{proj['ma_ct']}]: {e}")
        log(f"  ✗ Lỗi [{proj['ma_ct']}]: {e}")

log("=" * 55)
log(f"KẾT QUẢ:")
log(f"  Tạo mới : {created}")
log(f"  Cập nhật: {updated}")
log(f"  Lỗi     : {len(errors)}")
if errors:
    for e in errors: log(f"    {e}")
log("ĐỒNG BỘ CÔNG TRÌNH HOÀN TẤT")
log("=" * 55)

cur.close()
conn.close()
