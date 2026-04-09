"""
sync_bank_accounts.py
Import 31 tài khoản ngân hàng từ MISA vào Odoo 17
  - Tạo ngân hàng (res.bank) nếu chưa có
  - Tạo res.partner.bank liên kết với partner phù hợp
"""
import openpyxl, psycopg2, re, sys
sys.stdout.reconfigure(encoding='utf-8')

FILE = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_tai_khoan_ngan_hang.xlsx"

# BIC chuẩn của các ngân hàng Việt Nam (tra cứu)
BANK_BIC = {
    'Ngân hàng TMCP Ngoại thương Việt Nam':          'BFTVVNVX',   # Vietcombank
    'Ngân hàng TMCP Quốc Tế Việt Nam':               'VNIBVNVX',   # VIB
    'Ngân hàng TMCP Quân đội':                        'MSCBVNVX',   # MBBank
    'Ngân hàng TMCP An Bình':                         'ABBKVNVX',   # ABBank
    'Ngân hàng TMCP Đại Chúng Việt Nam':              'PVCOMVNVX',  # PVcomBank
    'Ngân hàng TMCP Kỹ thương Việt Nam':              'VTCBVNVX',   # Techcombank
    'Ngân hàng Việt Nam Thịnh Vượng':                 'VPBKVNVX',   # VPBank
    'Ngân hàng TMCP Á Châu':                          'ASCBVNVX',   # ACB
    'Ngân hàng TMCP Thịnh Vượng và Phát triển':       'PGBLVNVX',   # PGBank
    'Ngân hàng TMCP Đầu tư và Phát triển Việt Nam':   'BIDVVNVX',   # BIDV
}

conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()

# ── Load dữ liệu cần thiết ───────────────────────────────────────────────────
cur.execute("SELECT id FROM res_company ORDER BY id LIMIT 1;")
company_id = cur.fetchone()[0]

# Company partner (dùng cho TK của công ty)
cur.execute("SELECT partner_id FROM res_company WHERE id=%s;", (company_id,))
company_partner_id = cur.fetchone()[0]

# Load banks hiện có: name → id
cur.execute("SELECT id, name FROM res_bank;")
bank_map = {}
for r in cur.fetchall():
    bank_map[r[1].strip().lower()] = r[0]

# Load partners: name → id  (để khớp chủ tài khoản)
cur.execute("SELECT id, name FROM res_partner WHERE active=TRUE;")
partner_map = {}
for r in cur.fetchall():
    if r[1]:
        partner_map[r[1].strip().lower()] = r[0]

# Load số TK đã có
cur.execute("SELECT acc_number FROM res_partner_bank;")
existing_acc = {r[0] for r in cur.fetchall()}

print(f"Banks hiện có: {len(bank_map)}")
print(f"Partners hiện có: {len(partner_map)}")
print(f"TK ngân hàng đã có: {len(existing_acc)}")

# ── Đọc Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb.active

accounts = []
for r in range(4, ws.max_row + 1):
    so_tk   = str(ws.cell(row=r, column=2).value or '').strip()
    ten_nh  = str(ws.cell(row=r, column=3).value or '').strip()
    chi_nhanh = str(ws.cell(row=r, column=4).value or '').strip()
    chu_tk  = str(ws.cell(row=r, column=5).value or '').strip()
    trang_thai = str(ws.cell(row=r, column=7).value or '').strip()
    if not so_tk or trang_thai == 'Ngừng sử dụng':
        continue
    accounts.append({
        'so_tk': so_tk,
        'ten_nh': ten_nh.strip(),
        'chi_nhanh': chi_nhanh,
        'chu_tk': chu_tk,
    })

print(f"\nĐọc được {len(accounts)} tài khoản từ Excel\n")

# ── Hàm tìm hoặc tạo bank ────────────────────────────────────────────────────
def get_or_create_bank(ten_nh):
    key = ten_nh.strip().lower()
    if key in bank_map:
        return bank_map[key]
    bic = BANK_BIC.get(ten_nh.strip(), None)
    cur.execute("""
        INSERT INTO res_bank (name, bic, active, create_uid, write_uid, create_date, write_date)
        VALUES (%s, %s, TRUE, 2, 2, NOW(), NOW()) RETURNING id;
    """, (ten_nh.strip(), bic))
    bid = cur.fetchone()[0]
    bank_map[key] = bid
    print(f"  🏦 Tạo bank: {ten_nh} (BIC={bic}, id={bid})")
    return bid

# ── Hàm tìm partner ──────────────────────────────────────────────────────────
COMPANY_NAMES = {
    'công ty cổ phần thương mại vận tải hoàng sơn',
    'cty cp thương mại vận tải hoàng sơn',
    'công ty cp thương mại vận tải hoàng sơn',
}

def find_partner(chu_tk):
    if not chu_tk:
        return company_partner_id, None   # mặc định là công ty
    key = chu_tk.strip().lower()
    # Kiểm tra nếu là công ty chính
    if key in COMPANY_NAMES:
        return company_partner_id, None
    # Tìm trong partner_map (exact)
    if key in partner_map:
        return partner_map[key], None
    # Tìm partial match
    for pname, pid in partner_map.items():
        if key in pname or pname in key:
            return pid, None
    # Không tìm thấy → dùng company partner, ghi tên chủ TK vào acc_holder_name
    return company_partner_id, chu_tk.strip()

# ── Import từng tài khoản ────────────────────────────────────────────────────
created = 0
skipped = 0

for acc in accounts:
    so_tk = acc['so_tk']

    # Kiểm tra đã tồn tại chưa
    if so_tk in existing_acc:
        print(f"  ⏩ Đã có: {so_tk}")
        skipped += 1
        continue

    bank_id   = get_or_create_bank(acc['ten_nh'])
    partner_id, holder_name = find_partner(acc['chu_tk'])

    # sanitized = chỉ giữ chữ số và chữ cái
    sanitized = re.sub(r'[^A-Z0-9]', '', so_tk.upper())

    try:
        cur.execute("SAVEPOINT sp_bank;")
        cur.execute("""
            INSERT INTO res_partner_bank
                (acc_number, sanitized_acc_number, bank_id, partner_id,
                 acc_holder_name, active, company_id,
                 create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, %s, %s, %s, TRUE, %s, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (so_tk, sanitized, bank_id, partner_id,
              holder_name, company_id))
        new_id = cur.fetchone()[0]
        cur.execute("RELEASE SAVEPOINT sp_bank;")
        existing_acc.add(so_tk)
        created += 1

        holder_info = f" (chủ TK: {acc['chu_tk']})" if acc['chu_tk'] else ''
        bank_info   = f" | {acc['chi_nhanh']}" if acc['chi_nhanh'] else ''
        print(f"  ✅ {so_tk} — {acc['ten_nh']}{bank_info}{holder_info}")

    except Exception as e:
        cur.execute("ROLLBACK TO SAVEPOINT sp_bank;")
        print(f"  ❌ Lỗi {so_tk}: {e}")

conn.commit()
print(f"\n✅ HOÀN TẤT!")
print(f"   Tạo mới : {created} tài khoản")
print(f"   Bỏ qua  : {skipped} (đã có)")
cur.close()
conn.close()
