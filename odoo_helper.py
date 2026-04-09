"""
odoo_helper.py
--------------
Các hàm dùng chung cho sync_customers.py và sync_suppliers.py
"""
import re

TAG_FILE = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nhom_khach_hang_nha_cung_cap.xlsx"
EXT_MODULE = '__import__'

def get_or_create_tag(models, uid, db, password, ma_nhom, ten_nhom, tag_cache):
    """Lấy tag_id theo mã nhóm, tạo mới nếu chưa có. tag_cache là dict {ma: id}"""
    if ma_nhom in tag_cache:
        return tag_cache[ma_nhom]

    tag_name = f"{ma_nhom} - {ten_nhom}"
    existing = models.execute_kw(db, uid, password,
        'res.partner.category', 'search_read',
        [[['name', '=', tag_name]]], {'fields': ['id'], 'limit': 1})

    if existing:
        tag_id = existing[0]['id']
    else:
        tag_id = models.execute_kw(db, uid, password,
            'res.partner.category', 'create', [{'name': tag_name}])

    tag_cache[ma_nhom] = tag_id
    return tag_id

def load_all_tags(models, uid, db, password):
    """Load toàn bộ tags từ file nhóm MISA vào cache {ma_nhom: tag_id}"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(TAG_FILE, data_only=True)
        ws = wb.active
        tag_cache = {}
        colors = [1,2,3,4,5,6,7,8,9,10,11]
        i = 0
        for r in range(4, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, 4)]
            _, ma_nhom, ten_nhom = row
            if not ma_nhom or not ten_nhom:
                continue
            ma_nhom  = str(ma_nhom).strip()
            ten_nhom = str(ten_nhom).strip()
            tag_name = f"{ma_nhom} - {ten_nhom}"

            existing = models.execute_kw(db, uid, password,
                'res.partner.category', 'search_read',
                [[['name', '=', tag_name]]], {'fields': ['id'], 'limit': 1})
            if existing:
                tag_cache[ma_nhom] = existing[0]['id']
            else:
                tag_id = models.execute_kw(db, uid, password,
                    'res.partner.category', 'create',
                    [{'name': tag_name, 'color': colors[i % len(colors)]}])
                tag_cache[ma_nhom] = tag_id
            i += 1
        return tag_cache
    except Exception as e:
        print(f"  [WARN] Không load được tags: {e}")
        return {}

def assign_tag_to_partner(models, uid, db, password, partner_id, ma_nhom, tag_cache):
    """Gán tag cho partner nếu chưa có"""
    if ma_nhom not in tag_cache:
        return False
    tag_id = tag_cache[ma_nhom]

    partner = models.execute_kw(db, uid, password,
        'res.partner', 'read', [[partner_id]], {'fields': ['category_id']})[0]
    current = [t['id'] if isinstance(t, dict) else t for t in partner.get('category_id', [])]

    if tag_id not in current:
        models.execute_kw(db, uid, password,
            'res.partner', 'write',
            [[partner_id], {'category_id': [(4, tag_id)]}])
        return True
    return False
