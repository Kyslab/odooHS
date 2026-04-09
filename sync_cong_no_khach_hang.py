"""
sync_cong_no_khach_hang.py  (phiên bản chuẩn - có tra cứu Mã KH)
Nhập số dư công nợ đầu kỳ khách hàng từ MISA vào Odoo 17

Quy trình tìm/tạo partner:
  1. Tìm theo Mã KH (ref) đã có trong Odoo
  2. Tìm theo số điện thoại
  3. Tìm theo tên (chuẩn hóa)
  4. Nếu không tìm thấy → tạo mới, gán Mã KH từ file master
"""

import openpyxl, psycopg2, json, re, sys, logging, xmlrpc.client
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(filename=r'D:\odoo\sync_cong_no_khach_hang.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8')

# ── Đường dẫn file ────────────────────────────────────────────────────────────
CONG_NO_PATH  = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_cong_no_khach_hang.xlsx"
MASTER_PATH   = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khach_hang.xlsx"
OPENING_DATE  = '2024-12-31'
BALANCE_CODE  = '4111'
JOURNAL_CODE  = 'MISC'

ODOO_URL  = 'http://localhost:8017'
ODOO_DB   = 'odoo_company'
ODOO_UID  = 2
ODOO_PASS = 'admin'

def log(msg): print(msg); logging.info(msg)

def norm_phone(p):
    if not p: return ''
    s = re.sub(r'[^0-9]', '', str(p))
    if s.startswith('84') and len(s) >= 11: s = '0' + s[2:]
    return s

def norm_name(s):
    s = str(s).lower().strip()
    s = re.sub(r'\b0[0-9]{9,10}\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ─── 1. Load file master Mã KH ────────────────────────────────────────────────
log("📋 Load file master Mã KH...")
wb_m = openpyxl.load_workbook(MASTER_PATH, data_only=True)
ws_m = wb_m.active

master = {}   # ten_kh_norm → {'ma_kh', 'phone', 'mobile'}
master_by_ref = {}  # ma_kh → ten_kh

for r in range(4, ws_m.max_row + 1):
    ma_kh  = ws_m.cell(row=r, column=2).value
    ten_kh = ws_m.cell(row=r, column=3).value
    phone  = ws_m.cell(row=r, column=8).value
    mobile = ws_m.cell(row=r, column=10).value
    if not ma_kh or not ten_kh: continue
    ma_kh  = str(ma_kh).strip()
    ten_kh = str(ten_kh).strip()
    if ten_kh in ('#N/A','Tổng','') or ma_kh in ('0','#N/A','Tổng'): continue
    if not re.match(r'^[A-Za-z]+[0-9]+$', ma_kh): continue
    master[norm_name(ten_kh)] = {'ma_kh': ma_kh, 'phone': norm_phone(phone), 'mobile': norm_phone(mobile)}
    master_by_ref[ma_kh] = ten_kh

log(f"  Tải {len(master)} mã KH từ file master")

# ─── 2. Đọc file công nợ ──────────────────────────────────────────────────────
log("\n📂 Đọc file công nợ...")
wb = openpyxl.load_workbook(CONG_NO_PATH, data_only=True)
ws = wb.active

rows = []
for r in range(4, ws.max_row + 1):
    tk     = ws.cell(row=r, column=2).value
    ma_kh  = ws.cell(row=r, column=3).value
    ten_kh = ws.cell(row=r, column=4).value
    debit  = ws.cell(row=r, column=5).value or 0
    credit = ws.cell(row=r, column=6).value or 0
    if not tk or not ma_kh: continue
    if float(debit) == 0 and float(credit) == 0: continue
    rows.append({'tk': str(tk).strip(), 'ma_kh': str(ma_kh).strip(),
                 'ten_kh': str(ten_kh).strip() if ten_kh else '',
                 'debit': float(debit), 'credit': float(credit)})
log(f"  {len(rows)} dòng công nợ")

# ─── 3. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='localhost', port=5432, dbname=ODOO_DB,
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()
xm = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object', allow_none=True)

# Load partner lookup maps
cur.execute("SELECT id, name, ref, phone, mobile FROM res_partner WHERE active=TRUE;")
ref_map   = {}
phone_map = {}
name_map  = {}
for pid, pname, pref, pphone, pmobile in cur.fetchall():
    nm = list(pname.values())[0] if isinstance(pname, dict) else str(pname) if pname else ''
    if pref: ref_map[pref.strip()] = pid
    for ph in [pphone, pmobile]:
        np = norm_phone(ph)
        if np and len(np) >= 9: phone_map.setdefault(np, pid)
    nn = norm_name(nm)
    if nn: name_map.setdefault(nn, pid)

# Account & journal lookups
acc_codes = list({r['tk'] for r in rows})
cur.execute("SELECT code, id FROM account_account WHERE code = ANY(%s);", (acc_codes,))
acc_map = {r[0]: r[1] for r in cur.fetchall()}
cur.execute("SELECT id FROM account_account WHERE code=%s;", (BALANCE_CODE,))
balance_acc_id = cur.fetchone()[0]
cur.execute("SELECT id FROM account_journal WHERE code=%s;", (JOURNAL_CODE,))
journal_id = cur.fetchone()[0]

# ─── 4. Tìm / tạo partner ─────────────────────────────────────────────────────
log("\n👥 Tìm/tạo partner...")
partner_map = {}   # ma_kh → partner_id
unique_kh = {}
for r in rows:
    if r['ma_kh'] not in unique_kh:
        unique_kh[r['ma_kh']] = r['ten_kh']

stats = {'ref': 0, 'phone': 0, 'name': 0, 'new': 0}

for ma_kh, ten_kh in unique_kh.items():
    # 1. Tìm theo ref
    if ma_kh in ref_map:
        partner_map[ma_kh] = ref_map[ma_kh]
        stats['ref'] += 1
        continue

    # 2. Tìm theo SĐT (từ file master)
    found_id = None
    m = master.get(norm_name(ten_kh), {})
    for ph in [m.get('phone',''), m.get('mobile','')]:
        if ph and ph in phone_map:
            found_id = phone_map[ph]
            break

    if found_id:
        partner_map[ma_kh] = found_id
        stats['phone'] += 1
        # Gán ref nếu chưa có
        cur.execute("UPDATE res_partner SET ref=%s, write_date=NOW() WHERE id=%s AND (ref IS NULL OR ref='');",
                    (ma_kh, found_id))
        ref_map[ma_kh] = found_id
        continue

    # 3. Tìm theo tên
    nn = norm_name(ten_kh)
    if nn in name_map:
        found_id = name_map[nn]
        partner_map[ma_kh] = found_id
        stats['name'] += 1
        cur.execute("UPDATE res_partner SET ref=%s, write_date=NOW() WHERE id=%s AND (ref IS NULL OR ref='');",
                    (ma_kh, found_id))
        ref_map[ma_kh] = found_id
        continue

    # 4. Tạo mới — lấy Mã KH từ master (ma_kh đã là mã từ file công nợ)
    ten_chinh = master_by_ref.get(ma_kh, ten_kh)
    name_json = json.dumps({"vi_VN": ten_chinh, "en_US": ten_chinh}, ensure_ascii=False)
    m_info = master.get(norm_name(ten_chinh), {})
    cur.execute("""
        INSERT INTO res_partner
            (name, ref, phone, mobile, customer_rank, supplier_rank,
             active, is_company, create_uid, write_uid, create_date, write_date)
        VALUES (%s, %s, NULLIF(%s,''), NULLIF(%s,''), 1, 0,
                TRUE, FALSE, 2, 2, NOW(), NOW())
        RETURNING id;
    """, (name_json, ma_kh, m_info.get('phone') or None, m_info.get('mobile') or None))
    new_id = cur.fetchone()[0]
    partner_map[ma_kh] = new_id
    ref_map[ma_kh] = new_id
    stats['new'] += 1
    log(f"  ➕ Tạo mới: {ma_kh} - {ten_chinh}")

conn.commit()
log(f"  Khớp ref={stats['ref']} | SĐT={stats['phone']} | tên={stats['name']} | mới={stats['new']}")

# ─── 5. Tạo bút toán ──────────────────────────────────────────────────────────
log("\n📋 Tạo bút toán đầu kỳ...")
move_lines = []
total_debit = total_credit = 0.0

for r in rows:
    acc_id = acc_map.get(r['tk'])
    if not acc_id:
        log(f"  ⚠️  Không có TK {r['tk']}, bỏ qua")
        continue
    move_lines.append({
        'account_id': acc_id,
        'partner_id': partner_map.get(r['ma_kh']),
        'name'      : f"Số dư đầu kỳ {r['ma_kh']} - {r['ten_kh'][:50]}",
        'debit'     : r['debit'], 'credit': r['credit'],
    })
    total_debit += r['debit']; total_credit += r['credit']

diff = round(total_debit - total_credit, 2)
move_lines.append({
    'account_id': balance_acc_id,
    'name'      : 'Vốn chủ sở hữu đầu kỳ - công nợ KH (TK 4111)',
    'debit'     : 0.0 if diff > 0 else abs(diff),
    'credit'    : diff if diff > 0 else 0.0,
})

move_id = xm.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS, 'account.move', 'create', [{
    'journal_id': journal_id, 'date': OPENING_DATE,
    'ref'       : 'Số dư đầu kỳ công nợ khách hàng - nhập từ MISA (31/12/2024)',
    'move_type' : 'entry',
    'line_ids'  : [(0, 0, line) for line in move_lines],
}])
xm.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS, 'account.move', 'action_post', [[move_id]])

cur.close(); conn.close()
log(f"""
╔══════════════════════════════════════════════════════╗
║   HOÀN TẤT CÔNG NỢ ĐẦU KỲ KHÁCH HÀNG               ║
╠══════════════════════════════════════════════════════╣
║  Move ID     : {move_id:<36} ║
║  Số dòng     : {len(move_lines):<36} ║
║  Tổng Dư Nợ  : {total_debit:>20,.0f}               ║
║  Tổng Dư Có  : {total_credit:>20,.0f}               ║
╚══════════════════════════════════════════════════════╝
""")
