"""
fix_cong_no_khach_hang.py
1. Hủy và xóa bút toán ID=27 vừa tạo nhầm
2. Xóa 93 partner trùng tạo hôm nay
3. Khớp partner cũ theo tên → gán ref (Mã KH)
4. Tạo lại bút toán đầu kỳ với đúng partner
"""

import openpyxl, psycopg2, json, sys, logging, xmlrpc.client
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(filename=r'D:\odoo\fix_cong_no_khach_hang.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8')

EXCEL_PATH   = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_cong_no_khach_hang.xlsx"
OPENING_DATE = '2024-12-31'
BALANCE_CODE = '4111'
JOURNAL_CODE = 'MISC'
BAD_MOVE_ID  = 27   # bút toán cần xóa

ODOO_URL  = 'http://localhost:8017'
ODOO_DB   = 'odoo_company'
ODOO_UID  = 2
ODOO_PASS = 'admin'

def log(msg): print(msg); logging.info(msg)

# ─── 1. Đọc Excel ─────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active
rows = []
for r in range(4, ws.max_row + 1):
    tk     = ws.cell(row=r, column=2).value
    ma_kh  = ws.cell(row=r, column=3).value
    ten_kh = ws.cell(row=r, column=4).value
    debit  = ws.cell(row=r, column=5).value or 0
    credit = ws.cell(row=r, column=6).value or 0
    if not tk or not ma_kh: continue
    if float(debit) == 0 and float(credit) == 0: continue
    rows.append({'tk': str(tk).strip(), 'ma_kh': str(ma_kh).strip(),
                 'ten_kh': str(ten_kh).strip() if ten_kh else '',
                 'debit': float(debit), 'credit': float(credit)})
log(f"  {len(rows)} dòng công nợ")

# ─── 2. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='localhost', port=5432, dbname=ODOO_DB,
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()
xm = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object', allow_none=True)

# ─── 3. Hủy và xóa bút toán sai ──────────────────────────────────────────────
log(f"\n🗑️  Xóa bút toán ID={BAD_MOVE_ID} qua SQL...")
cur.execute("SELECT id FROM account_move WHERE id=%s;", (BAD_MOVE_ID,))
if cur.fetchone():
    # Xóa các reconcile liên quan trước
    cur.execute("""
        DELETE FROM account_partial_reconcile
        WHERE debit_move_id IN (SELECT id FROM account_move_line WHERE move_id=%s)
           OR credit_move_id IN (SELECT id FROM account_move_line WHERE move_id=%s);
    """, (BAD_MOVE_ID, BAD_MOVE_ID))
    cur.execute("DELETE FROM account_move_line WHERE move_id=%s;", (BAD_MOVE_ID,))
    cur.execute("DELETE FROM account_move WHERE id=%s;", (BAD_MOVE_ID,))
    conn.commit()
    log(f"  ✅ Đã xóa bút toán ID={BAD_MOVE_ID}")
else:
    log(f"  ℹ️  Bút toán ID={BAD_MOVE_ID} không tồn tại, bỏ qua")

# ─── 4. Xóa 93 partner trùng tạo hôm nay ─────────────────────────────────────
log("\n🗑️  Xóa partner trùng tạo hôm nay...")
cur.execute("""
    SELECT id, ref FROM res_partner
    WHERE ref IS NOT NULL AND create_date::date = CURRENT_DATE;
""")
dup_partners = cur.fetchall()
if dup_partners:
    dup_ids = [r[0] for r in dup_partners]
    cur.execute("DELETE FROM res_partner WHERE id = ANY(%s);", (dup_ids,))
    log(f"  ✅ Đã xóa {len(dup_ids)} partner trùng")
else:
    log("  ℹ️  Không có partner trùng cần xóa")
conn.commit()

# ─── 5. Lấy toàn bộ partner cũ, khớp theo tên ────────────────────────────────
log("\n🔗 Khớp partner cũ theo tên...")
cur.execute("""
    SELECT id, name FROM res_partner
    WHERE active=TRUE AND create_date::date < CURRENT_DATE;
""")
odoo_partners = cur.fetchall()

# Build map: tên chuẩn hóa → (id, tên gốc)
def norm(s):
    return str(s).lower().strip()

odoo_name_map = {}
for pid, pname in odoo_partners:
    # name có thể là plain string hoặc JSON
    try:
        if isinstance(pname, dict):
            names = list(pname.values())
        elif str(pname).startswith('{'):
            names = list(json.loads(pname).values())
        else:
            names = [str(pname)]
    except:
        names = [str(pname)]
    for n in names:
        odoo_name_map[norm(n)] = pid

# Gom unique partners từ file
unique_partners = {}
for r in rows:
    if r['ma_kh'] not in unique_partners:
        unique_partners[r['ma_kh']] = r['ten_kh']

partner_map = {}   # ma_kh → partner_id
not_found = []
matched_count = 0

for ma_kh, ten_kh in unique_partners.items():
    pid = odoo_name_map.get(norm(ten_kh))
    if pid:
        partner_map[ma_kh] = pid
        matched_count += 1
    else:
        not_found.append((ma_kh, ten_kh))

log(f"  Khớp được: {matched_count}/{len(unique_partners)}")
if not_found:
    log(f"  ⚠️  Không khớp ({len(not_found)} partner):")
    for ma, ten in not_found:
        log(f"       {ma}: {ten}")

# ─── 6. Gán ref (Mã KH) cho các partner đã khớp ─────────────────────────────
log("\n✏️  Gán Mã KH (ref) cho partner đã khớp...")
updated_ref = 0
for ma_kh, pid in partner_map.items():
    cur.execute("""
        UPDATE res_partner SET ref=%s, write_date=NOW(), write_uid=2
        WHERE id=%s AND (ref IS NULL OR ref='');
    """, (ma_kh, pid))
    if cur.rowcount > 0:
        updated_ref += 1
conn.commit()
log(f"  ✅ Đã gán ref cho {updated_ref} partner")

# ─── 7. Tạo mới partner cho những người không khớp (nếu có) ──────────────────
if not_found:
    log(f"\n  Tạo mới {len(not_found)} partner không tìm thấy...")
    for ma_kh, ten_kh in not_found:
        cur.execute("""
            INSERT INTO res_partner
                (name, ref, customer_rank, supplier_rank, active, is_company,
                 create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, 1, 0, TRUE, FALSE, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (ten_kh, ma_kh))
        pid = cur.fetchone()[0]
        partner_map[ma_kh] = pid
    conn.commit()
    log(f"  ✅ Đã tạo {len(not_found)} partner mới")

# ─── 8. Lấy account IDs ───────────────────────────────────────────────────────
acc_codes = list({r['tk'] for r in rows})
cur.execute("SELECT code, id FROM account_account WHERE code = ANY(%s);", (acc_codes,))
acc_map = {r[0]: r[1] for r in cur.fetchall()}

cur.execute("SELECT id FROM account_account WHERE code=%s;", (BALANCE_CODE,))
balance_acc_id = cur.fetchone()[0]
cur.execute("SELECT id FROM account_journal WHERE code=%s;", (JOURNAL_CODE,))
journal_id = cur.fetchone()[0]

# ─── 9. Xây dựng và tạo bút toán ─────────────────────────────────────────────
log("\n📋 Tạo bút toán đầu kỳ...")
move_lines = []
total_debit = total_credit = 0.0

for r in rows:
    acc_id = acc_map.get(r['tk'])
    if not acc_id:
        log(f"  ⚠️  Không có TK {r['tk']}, bỏ qua")
        continue
    move_lines.append({
        'account_id': acc_id,
        'partner_id': partner_map.get(r['ma_kh']),
        'name'      : f"Số dư đầu kỳ {r['ma_kh']} - {r['ten_kh'][:50]}",
        'debit'     : r['debit'],
        'credit'    : r['credit'],
    })
    total_debit  += r['debit']
    total_credit += r['credit']

diff = round(total_debit - total_credit, 2)
move_lines.append({
    'account_id': balance_acc_id,
    'name'      : 'Vốn chủ sở hữu đầu kỳ - công nợ KH (TK 4111)',
    'debit'     : 0.0 if diff > 0 else abs(diff),
    'credit'    : diff if diff > 0 else 0.0,
})

move_id = xm.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
    'account.move', 'create', [{
        'journal_id': journal_id,
        'date'      : OPENING_DATE,
        'ref'       : 'Số dư đầu kỳ công nợ khách hàng - nhập từ MISA (31/12/2024)',
        'move_type' : 'entry',
        'line_ids'  : [(0, 0, line) for line in move_lines],
    }])
xm.execute_kw(ODOO_DB, ODOO_UID, ODOO_PASS,
    'account.move', 'action_post', [[move_id]])

cur.close(); conn.close()

log(f"""
╔══════════════════════════════════════════════════════════╗
║      HOÀN TẤT - CÔNG NỢ ĐẦU KỲ KHÁCH HÀNG (FIX)        ║
╠══════════════════════════════════════════════════════════╣
║  Move ID mới      : {move_id:<35} ║
║  Partner khớp cũ  : {matched_count:<35} ║
║  Partner tạo mới  : {len(not_found):<35} ║
║  Tổng Dư Nợ       : {total_debit:>20,.0f}               ║
║  Tổng Dư Có       : {total_credit:>20,.0f}               ║
║  TK 4111 cân đối  : {abs(diff):>20,.0f}               ║
╚══════════════════════════════════════════════════════════╝
""")
