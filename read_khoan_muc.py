import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
path = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khoan_muc_chi_phi_.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for r in range(1, ws.max_row + 1):
        row_data = [str(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column+1)]
        print(f"  Row {r}: {row_data}")
