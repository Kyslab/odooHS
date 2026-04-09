"""
sync_ma_thong_ke.py
Nhập 241 mã thống kê từ MISA vào Odoo 17 (analytic plan riêng)
+ Bắt buộc chọn mã thống kê khi ghi nhận doanh thu (TK 5xx, 7xx)
"""

import openpyxl, psycopg2, json, sys, logging
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    filename=r'D:\odoo\sync_ma_thong_ke.log',
    filemode='w',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    encoding='utf-8'
)

EXCEL_PATH = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_ma_thong_ke.xlsx"
PLAN_NAME_VI = "Mã thống kê"
PLAN_NAME_EN = "Statistical Codes"

# Bắt buộc chọn mã thống kê khi nhập dòng tài khoản bắt đầu bằng '5' hoặc '7'
ACCOUNT_PREFIXES_MANDATORY = ['5', '7']

ODOO_URL  = 'http://localhost:8017'
ODOO_DB   = 'odoo_company'
ODOO_UID  = 2
ODOO_PASS = 'admin'

def log(msg):
    print(msg)
    logging.info(msg)

# ─── 1. Đọc Excel ────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

items = []  # list of {'code': str, 'name': str}
for r in range(4, ws.max_row + 1):
    code  = ws.cell(row=r, column=2).value   # Cột B: Mã thống kê
    name  = ws.cell(row=r, column=3).value   # Cột C: Tên thống kê
    state = ws.cell(row=r, column=5).value   # Cột E: Trạng thái
    if not code or not name:
        continue
    code  = str(code).strip()
    name  = str(name).strip()
    state = str(state).strip() if state else ''
    if state == 'Ngừng sử dụng':
        continue
    # Bỏ qua tên placeholder
    if name in ('...', '…', ''):
        name = code  # dùng mã làm tên nếu không có tên
    items.append({'code': code, 'name': name})

log(f"  → Đọc được {len(items)} mã thống kê đang sử dụng")

# ─── 2. Xây dựng bảng tra cứu parent theo prefix ─────────────────────────────
all_codes = {item['code'] for item in items}

def find_parent_code(code):
    """Tìm code cha = prefix dài nhất của code hiện tại có trong danh sách"""
    for length in range(len(code) - 1, 0, -1):
        prefix = code[:length]
        if prefix in all_codes:
            return prefix
    return None

parent_map = {}
for item in items:
    parent_map[item['code']] = find_parent_code(item['code'])

# ─── 3. Kết nối DB ────────────────────────────────────────────────────────────
conn = psycopg2.connect(
    host='localhost', port=5432, dbname=ODOO_DB,
    user='odoo17', password='odoo17pass'
)
conn.autocommit = False
cur = conn.cursor()

try:
    cur.execute("SELECT id FROM res_company ORDER BY id LIMIT 1;")
    company_id = cur.fetchone()[0]
    log(f"  Company ID: {company_id}")

    # ─── 4. Tạo hoặc lấy Analytic Plan "Mã thống kê" qua XML-RPC ─────────────
    import xmlrpc.client
    xmodels = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')

    cur.execute("SELECT id FROM account_analytic_plan WHERE name::text ILIKE %s;",
                (f'%{PLAN_NAME_VI}%',))
    row = cur.fetchone()
    if row:
        plan_id = row[0]
        log(f"  Plan đã tồn tại: ID={plan_id}")
    else:
        plan_id = xmodels.execute_kw(
            ODOO_DB, ODOO_UID, ODOO_PASS,
            'account.analytic.plan', 'create',
            [{'name': PLAN_NAME_VI, 'description': PLAN_NAME_EN}]
        )
        log(f"  ✅ Đã tạo plan mới qua ORM: ID={plan_id} — {PLAN_NAME_VI}")

    # ─── 5. Lấy danh sách mã đã có ───────────────────────────────────────────
    cur.execute("""
        SELECT code FROM account_analytic_account
        WHERE plan_id = %s AND code IS NOT NULL;
    """, (plan_id,))
    existing_codes = {r[0] for r in cur.fetchall()}
    log(f"  Đã có {len(existing_codes)} mã thống kê trong DB")

    # ─── 6. Lấy id các mã đã tồn tại ─────────────────────────────────────────
    code_to_id = {}
    if existing_codes:
        cur.execute("""
            SELECT code, id FROM account_analytic_account
            WHERE plan_id = %s AND code IS NOT NULL;
        """, (plan_id,))
        for r in cur.fetchall():
            code_to_id[r[0]] = r[1]

    # ─── 7. Import theo thứ tự từ ngắn → dài (cha trước con) ─────────────────
    items_sorted = sorted(items, key=lambda x: (len(x['code']), x['code']))

    created = 0
    skipped = 0

    for item in items_sorted:
        code = item['code']
        name = item['name']

        if code in existing_codes:
            skipped += 1
            if code not in code_to_id:
                cur.execute("SELECT id FROM account_analytic_account WHERE plan_id=%s AND code=%s;",
                            (plan_id, code))
                r = cur.fetchone()
                if r:
                    code_to_id[code] = r[0]
            continue

        name_json = json.dumps({"vi_VN": name, "en_US": name}, ensure_ascii=False)
        parent_code = parent_map.get(code)
        parent_id   = code_to_id.get(parent_code) if parent_code else None

        cur.execute("""
            INSERT INTO account_analytic_account
                (name, code, plan_id, root_plan_id, company_id, active,
                 create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, %s, %s, %s, TRUE, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (name_json, code, plan_id, plan_id, company_id))
        new_id = cur.fetchone()[0]
        code_to_id[code] = new_id
        created += 1

        if created % 50 == 0:
            log(f"    ... đã nhập {created} mã thống kê")

    log(f"\n  ✅ Kết quả: Tạo mới {created}, Bỏ qua (đã có) {skipped}")

    # ─── 8. Cập nhật parent_id nếu cột tồn tại ───────────────────────────────
    cur.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name='account_analytic_account' AND column_name='parent_id';
    """)
    has_parent_col = cur.fetchone() is not None
    log(f"  account_analytic_account.parent_id exists: {has_parent_col}")

    if has_parent_col:
        updated_parents = 0
        for item in items_sorted:
            code = item['code']
            parent_code = parent_map.get(code)
            if parent_code and code in code_to_id and parent_code in code_to_id:
                cur.execute("""
                    UPDATE account_analytic_account
                    SET parent_id = %s, write_date = NOW()
                    WHERE id = %s AND plan_id = %s;
                """, (code_to_id[parent_code], code_to_id[code], plan_id))
                updated_parents += 1
        log(f"  ✅ Đã cập nhật {updated_parents} quan hệ cha-con")
    else:
        log("  ℹ️  Không có cột parent_id — dữ liệu phẳng")

    # ─── 9. Cấu hình bắt buộc chọn mã thống kê (account_analytic_applicability) ─
    log("\n  ⚙️  Cấu hình bắt buộc nhập mã thống kê khi ghi nhận doanh thu...")
    for prefix in ACCOUNT_PREFIXES_MANDATORY:
        cur.execute("""
            SELECT id FROM account_analytic_applicability
            WHERE analytic_plan_id = %s
              AND business_domain = 'general'
              AND account_prefix = %s;
        """, (plan_id, prefix))
        existing = cur.fetchone()
        if existing:
            cur.execute("""
                UPDATE account_analytic_applicability
                SET applicability = 'mandatory', write_date = NOW(), write_uid = 2
                WHERE id = %s;
            """, (existing[0],))
            log(f"  ✅ Đã CẬP NHẬT → mandatory cho TK prefix '{prefix}'")
        else:
            cur.execute("""
                INSERT INTO account_analytic_applicability
                    (analytic_plan_id, company_id, business_domain, applicability,
                     account_prefix, create_uid, write_uid, create_date, write_date)
                VALUES (%s, %s, 'general', 'mandatory', %s, 2, 2, NOW(), NOW());
            """, (plan_id, company_id, prefix))
            log(f"  ✅ Đã TẠO quy tắc bắt buộc cho TK prefix '{prefix}'")

    conn.commit()
    log("\n✅ HOÀN TẤT! Đã commit tất cả thay đổi.")
    log(f"   Plan ID       : {plan_id}")
    log(f"   Tổng mã       : {created + skipped}")
    log(f"   Mới tạo       : {created}")
    log(f"   Đã có sẵn     : {skipped}")
    log(f"   Bắt buộc nhập : TK bắt đầu bằng {ACCOUNT_PREFIXES_MANDATORY}")

except Exception as e:
    conn.rollback()
    log(f"\n❌ LỖI: {e}")
    import traceback
    traceback.print_exc()
    raise
finally:
    cur.close()
    conn.close()
