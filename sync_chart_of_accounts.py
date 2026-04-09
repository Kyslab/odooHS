"""
sync_chart_of_accounts.py
So sánh hệ thống tài khoản MISA với Odoo, thêm tài khoản còn thiếu
"""
import openpyxl, psycopg2, json, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_he_thong_tai_khoan_.xlsx"

# ─── Mapping tính chất tài khoản MISA → account_type Odoo ─────────────────────
def get_account_type(code):
    c = str(code)
    # Tiền mặt / Ngân hàng / Tiền đang chuyển
    if c.startswith(('111','112','113')):
        return 'asset_cash', False
    # Phải thu khách hàng
    if c.startswith('131'):
        return 'asset_receivable', True
    # Chứng khoán, đầu tư ngắn hạn
    if c.startswith(('121','128')):
        return 'asset_current', False
    # Thuế GTGT được khấu trừ, các khoản phải thu ngắn hạn khác
    if c.startswith(('133','136','138','141','161','171')):
        return 'asset_current', False
    # Hàng tồn kho
    if c.startswith(('152','153','154','155','156','157','158')):
        return 'asset_current', False
    # Tài sản cố định hữu hình / vô hình / thuê tài chính
    if c.startswith(('211','212','213')):
        return 'asset_fixed', False
    # Hao mòn TSCĐ (tài khoản âm - contra asset)
    if c.startswith('214'):
        return 'asset_fixed', False
    # BĐS đầu tư, đầu tư dài hạn
    if c.startswith(('217','221','222','228','229')):
        return 'asset_non_current', False
    # XDCB dở dang, chi phí trả trước dài hạn
    if c.startswith(('241','242','243','244')):
        return 'asset_non_current', False
    # Toàn bộ tài sản 1xx, 2xx còn lại
    if c[0] in ('1','2'):
        if len(c) >= 3 and c[:3] in ('191','291'):
            return 'asset_non_current', False
        return 'asset_current', False
    # Phải trả nhà cung cấp
    if c.startswith('331'):
        return 'liability_payable', True
    # Thuế và các khoản phải nộp, phải trả ngắn hạn
    if c.startswith(('311','315','333','334','335','336','337','338','341')):
        return 'liability_current', False
    # Vay dài hạn, trái phiếu, nợ dài hạn
    if c.startswith(('341','343','344','347','352','353','356','357')):
        return 'liability_non_current', False
    if c[0] == '3':
        return 'liability_current', False
    # Vốn chủ sở hữu
    if c.startswith('421'):
        return 'equity_unaffected', False
    if c[0] == '4':
        return 'equity', False
    # Doanh thu
    if c.startswith('511') or c.startswith('512'):
        return 'income', False
    if c.startswith('515'):
        return 'income', False
    if c.startswith('521'):
        return 'income', False   # Các khoản giảm trừ DT
    if c.startswith('711'):
        return 'income_other', False
    # Chi phí giá vốn
    if c.startswith('632') or c.startswith('631'):
        return 'expense_direct_cost', False
    # Khấu hao
    if c.startswith(('6274','6234','6414','6424')):
        return 'expense_depreciation', False
    # Chi phí tài chính, bán hàng, QLDN, khác
    if c[0] in ('6','7','8'):
        return 'expense', False
    # Ngoài bảng
    if c[0] == '0':
        return 'off_balance', False
    return 'asset_current', False

# ─── 1. Đọc MISA ──────────────────────────────────────────────────────────────
print("📂 Đọc file MISA...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

misa_accounts = {}   # code → {name_vi, name_en, nature}
for r in range(4, ws.max_row + 1):
    code    = ws.cell(row=r, column=2).value
    name_vi = ws.cell(row=r, column=3).value
    nature  = ws.cell(row=r, column=4).value
    name_en = ws.cell(row=r, column=5).value
    if not code or not name_vi:
        continue
    code = str(code).strip()
    misa_accounts[code] = {
        'name_vi': str(name_vi).strip(),
        'name_en': str(name_en).strip() if name_en else str(name_vi).strip(),
        'nature':  str(nature).strip() if nature else '',
    }

print(f"  → Đọc được {len(misa_accounts)} tài khoản từ MISA")

# ─── 2. Đọc Odoo ──────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='localhost', port=5432, dbname='odoo_company',
                        user='odoo17', password='odoo17pass')
conn.autocommit = False
cur = conn.cursor()

cur.execute("SELECT code FROM account_account WHERE company_id=1;")
odoo_codes = {r[0] for r in cur.fetchall()}
print(f"  → Hiện có {len(odoo_codes)} tài khoản trong Odoo")

cur.execute("SELECT id FROM res_company ORDER BY id LIMIT 1;")
company_id = cur.fetchone()[0]

# ─── 3. Tìm tài khoản còn thiếu ──────────────────────────────────────────────
missing = {code: data for code, data in misa_accounts.items()
           if code not in odoo_codes}

print(f"\n📋 Tài khoản có trong MISA nhưng chưa có trong Odoo: {len(missing)}")
for code in sorted(missing.keys()):
    print(f"  {code:<12} {missing[code]['name_vi']}")

# ─── 4. Hỏi xác nhận trước khi insert ────────────────────────────────────────
print(f"\n➡️  Sẽ thêm {len(missing)} tài khoản mới vào Odoo.")
print("Bắt đầu import...")

# ─── 5. Load account_root map (code prefix 2 ký tự → root id) ────────────────
cur.execute("SELECT id, name FROM account_root WHERE company_id=%s;", (company_id,))
root_map = {r[1]: r[0] for r in cur.fetchall()}

def get_root_id(code):
    """Tìm account_root.id phù hợp với code: thử 2 ký tự trước, rồi 1 ký tự"""
    return root_map.get(code[:2]) or root_map.get(code[:1])

# ─── 6. Import tài khoản còn thiếu ───────────────────────────────────────────
created = 0
errors  = 0

# Sắp xếp theo code để cha trước con
for code in sorted(missing.keys()):
    data = missing[code]
    name_json = json.dumps({"vi_VN": data['name_vi'], "en_US": data['name_en']},
                            ensure_ascii=False)
    acc_type, reconcile = get_account_type(code)

    # Xác định include_initial_balance dựa theo nhóm tài khoản
    code1 = code[0]
    include_initial = code1 in ('1','2','3','4')   # BS accounts

    # internal_group theo account_type
    igroup_map = {
        'asset_receivable':'asset','asset_cash':'asset','asset_current':'asset',
        'asset_non_current':'asset','asset_prepayments':'asset','asset_fixed':'asset',
        'liability_payable':'liability','liability_credit_card':'liability',
        'liability_current':'liability','liability_non_current':'liability',
        'equity':'equity','equity_unaffected':'equity',
        'income':'income','income_other':'income',
        'expense':'expense','expense_depreciation':'expense','expense_direct_cost':'expense',
        'off_balance':'off_balance',
    }
    internal_group = igroup_map.get(acc_type, 'asset')

    root_id = get_root_id(code)

    try:
        cur.execute("""
            INSERT INTO account_account
                (code, name, account_type, internal_group, reconcile, deprecated,
                 include_initial_balance, non_trade, root_id,
                 company_id, create_uid, write_uid, create_date, write_date)
            VALUES (%s, %s, %s, %s, %s, FALSE, %s, FALSE, %s,
                    %s, 2, 2, NOW(), NOW())
            RETURNING id;
        """, (code, name_json, acc_type, internal_group, reconcile, include_initial, root_id, company_id))

        result = cur.fetchone()
        if result:
            created += 1
            if created % 20 == 0:
                print(f"    ... đã tạo {created} tài khoản")

    except Exception as e:
        print(f"  ❌ Lỗi TK {code}: {e}")
        conn.rollback()
        errors += 1
        continue

try:
    conn.commit()
    print(f"\n✅ HOÀN TẤT!")
    print(f"   Tạo mới:  {created}")
    print(f"   Lỗi:      {errors}")
except Exception as e:
    conn.rollback()
    print(f"❌ Commit lỗi: {e}")
finally:
    cur.close()
    conn.close()
