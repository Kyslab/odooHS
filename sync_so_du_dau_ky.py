"""
sync_so_du_dau_ky.py
Tạo bút toán số dư đầu kỳ từ file MISA vào Odoo 17
- Chỉ dùng tài khoản LÁ (không dùng tài khoản tổng hợp cha) để tránh tính 2 lần
- Tài khoản cân đối: TK 4111 (Vốn đầu tư của chủ sở hữu)
- Ngày: 31/12/2024 (ngày cuối kỳ trước khi mở sổ Odoo)
"""

import openpyxl, psycopg2, json, sys, logging, xmlrpc.client
from datetime import date
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    filename=r'D:\odoo\sync_so_du_dau_ky.log',
    filemode='w', level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s', encoding='utf-8'
)

EXCEL_PATH    = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_so_du_tai_khoan.xlsx"
OPENING_DATE  = '2024-12-31'   # Ngày số dư đầu kỳ
BALANCE_CODE  = '4111'         # TK cân đối: Vốn đầu tư của chủ sở hữu
JOURNAL_CODE  = 'MISC'         # Nhật ký sử dụng

ODOO_URL  = 'http://localhost:8017'
ODOO_DB   = 'odoo_company'
ODOO_UID  = 2
ODOO_PASS = 'admin'

def log(msg):
    print(msg); logging.info(msg)

# ─── 1. Đọc Excel ─────────────────────────────────────────────────────────────
log("📂 Đọc file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

raw = []
for r in range(4, ws.max_row + 1):
    code  = ws.cell(row=r, column=2).value
    name  = ws.cell(row=r, column=3).value
    debit = ws.cell(row=r, column=4).value or 0
    credit= ws.cell(row=r, column=5).value or 0
    if not code: continue
    code = str(code).strip()
    raw.append({'code': code, 'name': str(name).strip() if name else code,
                'debit': float(debit), 'credit': float(credit)})

log(f"  Đọc được {len(raw)} dòng")

# ─── 2. Xác định tài khoản LÁ (loại bỏ tài khoản cha tổng hợp) ───────────────
all_codes = {item['code'] for item in raw}

def is_parent(code):
    """True nếu code là tiền tố của ít nhất 1 code khác trong file"""
    return any(other != code and other.startswith(code) for other in all_codes)

leaves = [item for item in raw if not is_parent(item['code'])]
parents_skipped = [item['code'] for item in raw if is_parent(item['code'])]

log(f"\n  Tài khoản lá (dùng để nhập): {len(leaves)}")
log(f"  Tài khoản cha (bỏ qua - tổng hợp): {parents_skipped}")

# ─── 3. Kết nối Odoo DB lấy account_id cho từng mã ───────────────────────────
log("\n🔗 Kết nối DB...")
conn = psycopg2.connect(host='localhost', port=5432, dbname=ODOO_DB,
                         user='odoo17', password='odoo17pass')
conn.autocommit = True
cur = conn.cursor()

leaf_codes = [item['code'] for item in leaves]
cur.execute("SELECT code, id FROM account_account WHERE code = ANY(%s);", (leaf_codes,))
code_to_id = {r[0]: r[1] for r in cur.fetchall()}

missing = [c for c in leaf_codes if c not in code_to_id]
if missing:
    log(f"  ⚠️  Các mã KHÔNG TỒN TẠI trong Odoo: {missing}")

# Lấy journal_id
cur.execute("SELECT id FROM account_journal WHERE code = %s;", (JOURNAL_CODE,))
journal_row = cur.fetchone()
if not journal_row:
    raise ValueError(f"Không tìm thấy journal '{JOURNAL_CODE}'")
journal_id = journal_row[0]
log(f"  Journal '{JOURNAL_CODE}': ID={journal_id}")

# Lấy TK cân đối 4111
cur.execute("SELECT id FROM account_account WHERE code = %s;", (BALANCE_CODE,))
balance_row = cur.fetchone()
if not balance_row:
    raise ValueError(f"Không tìm thấy tài khoản '{BALANCE_CODE}'")
balance_account_id = balance_row[0]
log(f"  TK cân đối '{BALANCE_CODE}': ID={balance_account_id}")

cur.execute("SELECT id FROM res_company ORDER BY id LIMIT 1;")
company_id = cur.fetchone()[0]
cur.close(); conn.close()

# ─── 4. Xây dựng các dòng journal entry ──────────────────────────────────────
log("\n📋 Xây dựng dòng bút toán...")
move_lines = []
total_debit  = 0.0
total_credit = 0.0

for item in leaves:
    code = item['code']
    if code not in code_to_id:
        log(f"  ⚠️  Bỏ qua (không có trong Odoo): {code}")
        continue

    acc_id = code_to_id[code]

    if item['debit'] > 0:
        move_lines.append({
            'account_id': acc_id,
            'name'      : f"Số dư đầu kỳ {code} - {item['name']}",
            'debit'     : item['debit'],
            'credit'    : 0.0,
        })
        total_debit += item['debit']
        log(f"  Nợ  {code}: {item['debit']:>20,.0f}")

    if item['credit'] > 0:
        move_lines.append({
            'account_id': acc_id,
            'name'      : f"Số dư đầu kỳ {code} - {item['name']}",
            'debit'     : 0.0,
            'credit'    : item['credit'],
        })
        total_credit += item['credit']
        log(f"  Có  {code}: {item['credit']:>20,.0f}")

log(f"\n  Tổng Nợ   : {total_debit:>20,.0f}")
log(f"  Tổng Có   : {total_credit:>20,.0f}")

# ─── 5. Thêm dòng cân đối vào TK 4111 ────────────────────────────────────────
diff = round(total_debit - total_credit, 2)
log(f"  Chênh lệch: {diff:>20,.0f}")

if diff > 0:
    # Tổng Nợ > Tổng Có → cần Có thêm vào TK 4111
    move_lines.append({
        'account_id': balance_account_id,
        'name'      : 'Vốn chủ sở hữu đầu kỳ (TK 4111)',
        'debit'     : 0.0,
        'credit'    : diff,
    })
    log(f"  ➕ Cân đối: Có TK {BALANCE_CODE} = {diff:,.0f}")
elif diff < 0:
    # Tổng Có > Tổng Nợ → cần Nợ thêm vào TK 4111
    move_lines.append({
        'account_id': balance_account_id,
        'name'      : 'Vốn chủ sở hữu đầu kỳ (TK 4111)',
        'debit'     : abs(diff),
        'credit'    : 0.0,
    })
    log(f"  ➕ Cân đối: Nợ TK {BALANCE_CODE} = {abs(diff):,.0f}")
else:
    log("  ✅ Bút toán đã cân bằng, không cần dòng cân đối")

# ─── 6. Tạo bút toán qua XML-RPC ─────────────────────────────────────────────
log("\n🚀 Tạo bút toán số dư đầu kỳ qua XML-RPC...")
xmodels = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')

move_vals = {
    'journal_id' : journal_id,
    'date'       : OPENING_DATE,
    'ref'        : 'Số dư đầu kỳ - nhập từ MISA (31/12/2024)',
    'move_type'  : 'entry',
    'line_ids'   : [(0, 0, line) for line in move_lines],
}

move_id = xmodels.execute_kw(
    ODOO_DB, ODOO_UID, ODOO_PASS,
    'account.move', 'create', [move_vals]
)
log(f"  ✅ Đã tạo bút toán ID={move_id} (trạng thái: Nháp)")

# ─── 7. Xác nhận (post) bút toán ─────────────────────────────────────────────
log("  📌 Đang xác nhận (post) bút toán...")
xmodels.execute_kw(
    ODOO_DB, ODOO_UID, ODOO_PASS,
    'account.move', 'action_post', [[move_id]]
)
log(f"  ✅ Bút toán ID={move_id} đã được XÁC NHẬN (Posted)")

log(f"""
╔══════════════════════════════════════════════════════╗
║         HOÀN TẤT NHẬP SỐ DƯ ĐẦU KỲ                 ║
╠══════════════════════════════════════════════════════╣
║  Move ID      : {move_id:<36} ║
║  Ngày         : {OPENING_DATE:<36} ║
║  Nhật ký      : {JOURNAL_CODE:<36} ║
║  Số dòng      : {len(move_lines):<36} ║
║  Tổng Nợ      : {total_debit:>20,.0f}               ║
║  Tổng Có      : {total_credit:>20,.0f}               ║
║  TK cân đối   : {BALANCE_CODE} = {diff:,.0f}
╚══════════════════════════════════════════════════════╝
""")
