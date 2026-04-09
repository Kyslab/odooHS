# -*- coding: utf-8 -*-
"""
Gán mã nhân viên (LX/NV) từ file Excel vào Odoo.

Các trường hợp xử lý:
1. Partner HR chưa có ref, không có journal lines → gán ref trực tiếp
2. Partner HR chưa có ref, nhưng partner khác (kế toán) đã có ref + có lines
   → cập nhật hr_employee.work_contact_id sang partner kế toán, vô hiệu hoá partner HR
3. Partner HR chưa có ref, nhưng có partner trùng tên (dup, không có lines)
   → gán ref vào partner HR, vô hiệu hoá partner dup
4. Partner HR đã có ref đúng → bỏ qua
"""

import re
import sys
import psycopg2
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE = r'C:\Users\DELL\Downloads\in xong xoa\Danh_sach_nhan_vien.xlsx'
DB_CONN = dict(host='localhost', port=5432, dbname='odoo_company',
               user='odoo17', password='odoo17pass')

# ── helpers ────────────────────────────────────────────────────────────────────
def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').lower().strip())


# ── đọc Excel ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active

employees_xl = []   # [(ma, ten)]
for r in range(4, ws.max_row + 1):
    ma  = ws.cell(row=r, column=2).value
    ten = ws.cell(row=r, column=3).value
    if not ma or not ten:
        continue
    employees_xl.append((str(ma).strip(), str(ten).strip()))

print(f'Đọc từ Excel: {len(employees_xl)} nhân viên')

# ── kết nối DB ────────────────────────────────────────────────────────────────
conn = psycopg2.connect(**DB_CONN)
cur  = conn.cursor()


def get_partner(pid):
    cur.execute('''
        SELECT id, name::text, ref, supplier_rank, customer_rank, active,
               (SELECT COUNT(*) FROM account_move_line WHERE partner_id = res_partner.id)
        FROM res_partner WHERE id = %s
    ''', (pid,))
    return cur.fetchone()   # id, name, ref, sup, cust, active, lines


# ── đọc toàn bộ hr_employee ────────────────────────────────────────────────────
cur.execute('''
    SELECT e.id, e.name, e.work_contact_id
    FROM hr_employee e
    ORDER BY e.id
''')
emp_rows = cur.fetchall()   # [(emp_id, emp_name, work_contact_id)]

# map norm(name) → emp info
emp_by_name = {}
for eid, ename, pid in emp_rows:
    emp_by_name[norm(ename)] = (eid, ename, pid)

# ── tìm partner dup (cùng tên, khác partner HR) ────────────────────────────────
def find_dup_partner(hr_pid, emp_name):
    """Tìm partner khác cùng tên với partner HR."""
    cur.execute('''
        SELECT p.id, p.ref, p.supplier_rank, p.customer_rank,
               (SELECT COUNT(*) FROM account_move_line WHERE partner_id = p.id)
        FROM res_partner p
        WHERE p.active = TRUE
          AND p.id <> %s
          AND p.name::text = %s
    ''', (hr_pid, emp_name))
    return cur.fetchall()   # list of (id, ref, sup, cust, lines)


# ── xử lý từng nhân viên ──────────────────────────────────────────────────────
assigned = 0
skipped  = 0
updated_wc = 0
deactivated = 0
not_found = []

for ma, ten in employees_xl:
    key = norm(ten)
    emp = emp_by_name.get(key)
    if not emp:
        not_found.append((ma, ten))
        continue

    emp_id, emp_name, hr_pid = emp
    if not hr_pid:
        print(f'  [WARN] {ma} {ten}: không có work_contact_id')
        continue

    p = get_partner(hr_pid)
    if not p:
        print(f'  [WARN] {ma} {ten}: partner {hr_pid} không tồn tại')
        continue

    _, pname, cur_ref, sup, cust, active, lines = p

    # Trường hợp đã có đúng ref → bỏ qua
    if cur_ref == ma:
        print(f'  [OK]   {ma} {ten}: đã có ref (partner {hr_pid})')
        skipped += 1
        continue

    # Tìm partner dup cùng tên
    dups = find_dup_partner(hr_pid, ten)

    # Kiểm tra xem có dup nào đã có ref đúng + có journal lines không
    acc_partner = None
    for dp_id, dp_ref, dp_sup, dp_cust, dp_lines in dups:
        if dp_ref == ma and dp_lines > 0:
            acc_partner = (dp_id, dp_ref, dp_sup, dp_cust, dp_lines)
            break

    if acc_partner:
        # Case 2: Partner kế toán đã có ref + lines → chuyển work_contact_id
        dp_id, dp_ref, dp_sup, dp_cust, dp_lines = acc_partner
        cur.execute('UPDATE hr_employee SET work_contact_id=%s WHERE id=%s', (dp_id, emp_id))
        cur.execute('UPDATE res_partner SET active=FALSE WHERE id=%s', (hr_pid,))
        print(f'  [WC]   {ma} {ten}: chuyển EMP{emp_id} → partner {dp_id} (có {dp_lines} lines), vô hiệu partner {hr_pid}')
        updated_wc  += 1
        deactivated += 1

    else:
        # Case 1 & 3: Gán ref vào partner HR
        # Nếu có dup không có lines → lấy sup/cust từ dup rồi deactivate dup
        for dp_id, dp_ref, dp_sup, dp_cust, dp_lines in dups:
            if dp_lines == 0 and dp_ref is None:
                # sao chép supplier/customer rank nếu cao hơn
                new_sup  = max(sup,  dp_sup)
                new_cust = max(cust, dp_cust)
                cur.execute('UPDATE res_partner SET supplier_rank=%s, customer_rank=%s WHERE id=%s',
                            (new_sup, new_cust, hr_pid))
                cur.execute('UPDATE res_partner SET active=FALSE WHERE id=%s', (dp_id,))
                print(f'  [DUP]  {ma} {ten}: vô hiệu partner dup {dp_id} (sup={dp_sup},cust={dp_cust})')
                deactivated += 1
                sup, cust = new_sup, new_cust

        cur.execute('UPDATE res_partner SET ref=%s WHERE id=%s', (ma, hr_pid))
        print(f'  [SET]  {ma} {ten}: gán ref={ma} → partner {hr_pid}')
        assigned += 1

conn.commit()
cur.close()
conn.close()

print()
print('═' * 55)
print(f'Gán ref mới       : {assigned}')
print(f'Đã có ref (skip)  : {skipped}')
print(f'Chuyển work_contact: {updated_wc}')
print(f'Vô hiệu partner dup: {deactivated}')
if not_found:
    print(f'Không tìm thấy trong Odoo: {len(not_found)}')
    for ma, ten in not_found:
        print(f'   {ma} | {ten}')
