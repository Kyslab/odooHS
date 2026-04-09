import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Khach_hang"

# ---- Màu sắc ----
green  = PatternFill("solid", fgColor="1F7A4A")   # header chính
yellow = PatternFill("solid", fgColor="FFF2CC")   # bắt buộc
blue   = PatternFill("solid", fgColor="DDEEFF")   # tùy chọn
gray   = PatternFill("solid", fgColor="F2F2F2")   # ghi chú
white_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
black_font = Font(color="000000", name="Arial", size=10)
bold_font  = Font(color="000000", bold=True, name="Arial", size=10)
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

# ---- Cột & tiêu đề ----
# (odoo_field, header_vn, ghi_chu, bat_buoc, width)
columns = [
    ("External ID",          "Mã định danh (ID)",        "Ví dụ: KH001, KH002... (không trùng)",    True,  18),
    ("Name",                 "Tên khách hàng *",          "Bắt buộc",                                True,  35),
    ("Company Type",         "Loại",                      "company = Công ty  |  person = Cá nhân",  True,  14),
    ("Street",               "Địa chỉ",                   "",                                        False, 35),
    ("City",                 "Thành phố / Tỉnh",          "",                                        False, 20),
    ("Phone",                "Điện thoại",                "Ví dụ: 0901234567",                       False, 18),
    ("Mobile",               "Di động",                   "",                                        False, 18),
    ("Email",                "Email",                     "",                                        False, 25),
    ("VAT",                  "Mã số thuế (MST)",          "10 hoặc 13 chữ số",                       False, 18),
    ("Customer Rank",        "Là khách hàng",             "1 = có  |  0 = không",                    True,  16),
    ("Supplier Rank",        "Là nhà cung cấp",           "1 = có  |  0 = không",                    True,  16),
    ("Notes",                "Ghi chú",                   "",                                        False, 30),
]

# ---- Dòng 1: Hướng dẫn ----
ws.merge_cells("A1:L1")
ws["A1"] = "TEMPLATE IMPORT KHÁCH HÀNG / NHÀ CUNG CẤP — Odoo 17"
ws["A1"].font = Font(color="FFFFFF", bold=True, name="Arial", size=12)
ws["A1"].fill = green
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ---- Dòng 2: Ghi chú màu ----
ws.merge_cells("A2:L2")
ws["A2"] = "⚠  Cột nền VÀNG = bắt buộc nhập  |  Cột nền XANH = tùy chọn  |  Không xóa dòng 3 (tên cột Odoo)"
ws["A2"].font = Font(color="7B3F00", bold=True, name="Arial", size=10)
ws["A2"].fill = PatternFill("solid", fgColor="FFF2CC")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# ---- Dòng 3: Tên field Odoo (để Odoo nhận dạng) ----
for col_idx, (field, _, _, required, width) in enumerate(columns, 1):
    cell = ws.cell(row=3, column=col_idx, value=field)
    cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=9)
    cell.fill = PatternFill("solid", fgColor="2E4057")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[3].height = 20

# ---- Dòng 4: Tên tiếng Việt ----
for col_idx, (_, header_vn, ghi_chu, required, _) in enumerate(columns, 1):
    cell = ws.cell(row=4, column=col_idx, value=header_vn)
    cell.font = bold_font
    cell.fill = yellow if required else blue
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin
ws.row_dimensions[4].height = 22

# ---- Dòng 5: Ghi chú ----
for col_idx, (_, _, ghi_chu, _, _) in enumerate(columns, 1):
    cell = ws.cell(row=5, column=col_idx, value=ghi_chu)
    cell.font = Font(color="666666", italic=True, name="Arial", size=9)
    cell.fill = gray
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin
ws.row_dimensions[5].height = 30

# ---- Dữ liệu mẫu (dòng 6-10) ----
samples = [
    ["KH001", "Công ty TNHH ABC",         "company", "123 Nguyễn Văn Cừ, Q.5", "TP Hồ Chí Minh", "028 3812 3456", "0901 234 567", "abc@company.com",    "0123456789",  1, 0, "Khách lâu năm"],
    ["KH002", "Công ty CP XYZ",           "company", "45 Trần Duy Hưng, Cầu Giấy", "Hà Nội",      "024 3556 7890", "0912 345 678", "xyz@xyz.vn",         "9876543210",  1, 0, ""],
    ["KH003", "Nguyễn Văn An",            "person",  "78 Lê Lợi, Q.1",            "TP Hồ Chí Minh","",            "0933 456 789", "nguyenvanan@gmail.com","",           1, 0, "Khách lẻ"],
    ["NCC001","Công ty TNHH Vật tư ABC",  "company", "10 Công Nghiệp, Bình Dương", "Bình Dương",   "0274 382 1234", "0988 111 222", "vattu@abc.vn",       "0246813579",  0, 1, "Nhà cung cấp chính"],
    ["KH004", "Trần Thị Bích",            "person",  "22 Pasteur, Q.3",            "TP Hồ Chí Minh","",            "0977 654 321", "",                   "",            1, 0, ""],
]

for r, row_data in enumerate(samples, 6):
    for c, value in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = black_font
        cell.alignment = Alignment(vertical="center")
        cell.border = thin
        # highlight bắt buộc
        if c in [1, 2, 3, 10, 11]:
            cell.fill = PatternFill("solid", fgColor="FFFDE7")
    ws.row_dimensions[r].height = 18

# ---- Sheet 2: Hướng dẫn ----
ws2 = wb.create_sheet("Huong_dan")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 60

guide = [
    ["BƯỚC 1 — EXPORT TỪ MISA AMIS", ""],
    ["", "1. Đăng nhập MISA AMIS"],
    ["", "2. Vào: Danh mục → Khách hàng → Xuất Excel"],
    ["", "3. Tương tự cho Nhà cung cấp: Danh mục → Nhà cung cấp → Xuất Excel"],
    ["", ""],
    ["BƯỚC 2 — COPY DỮ LIỆU VÀO FILE NÀY", ""],
    ["", "Copy từng cột MISA sang cột tương ứng trong sheet 'Khach_hang'"],
    ["", "XÓA các dòng mẫu (dòng 6-10) trước khi dán dữ liệu thật"],
    ["", ""],
    ["BƯỚC 3 — MAP CỘT MISA → ODOO", ""],
    ["MISA AMIS", "Odoo (cột trong file này)"],
    ["Mã khách hàng",   "External ID  (ví dụ: KH001)"],
    ["Tên khách hàng",  "Name"],
    ["Loại (CTY/CN)",   "Company Type: company hoặc person"],
    ["Địa chỉ",         "Street"],
    ["Tỉnh/Thành phố",  "City"],
    ["Điện thoại",      "Phone"],
    ["Di động",         "Mobile"],
    ["Email",           "Email"],
    ["Mã số thuế",      "VAT"],
    ["(là khách hàng)", "Customer Rank = 1"],
    ["(là nhà cung cấp)","Supplier Rank = 1"],
    ["", ""],
    ["BƯỚC 4 — IMPORT VÀO ODOO", ""],
    ["", "1. Vào Odoo: Invoicing → Customers → Action → Import Records"],
    ["", "2. Upload file Excel này"],
    ["", "3. Odoo tự nhận cột từ dòng 3 (tên field tiếng Anh)"],
    ["", "4. Nhấn Test Import trước để kiểm tra lỗi"],
    ["", "5. Nhấn Import để hoàn tất"],
    ["", ""],
    ["LƯU Ý QUAN TRỌNG", ""],
    ["", "- External ID phải KHÔNG trùng nhau"],
    ["", "- Company Type phải gõ đúng: 'company' hoặc 'person' (chữ thường)"],
    ["", "- Customer Rank và Supplier Rank chỉ điền 0 hoặc 1"],
    ["", "- MST: chỉ nhập số, không có dấu gạch ngang"],
    ["", "- Có thể để trống các cột không bắt buộc"],
]

for r, (a, b) in enumerate(guide, 1):
    ws2.cell(row=r, column=1, value=a).font = bold_font if a else black_font
    ws2.cell(row=r, column=2, value=b).font = black_font
    if a and not b:
        ws2.cell(row=r, column=1).fill = green
        ws2.cell(row=r, column=1).font = white_font

wb.save("D:/odoo/customers_import.xlsx")
print("Da tao file: D:/odoo/customers_import.xlsx")
