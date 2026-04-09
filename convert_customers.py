import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, re
sys.stdout.reconfigure(encoding='utf-8')

src = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khach_hang.xlsx"
wb_src = openpyxl.load_workbook(src, data_only=True)
ws_src = wb_src['Danh sách khách hàng']

# --- Đọc dữ liệu từ row 4 trở đi (row 3 là header, row 403 là tổng) ---
records = []
skipped = []

for r in range(4, ws_src.max_row):  # bỏ dòng tổng cuối
    row = [ws_src.cell(row=r, column=c).value for c in range(1, 12)]
    stt, ma_kh, ten_kh, dia_chi, cong_no, nhom, mst, dien_thoai, cmnd, di_dong, la_ncc = row

    # Bỏ qua dòng rỗng hoặc lỗi
    if not ten_kh or str(ten_kh).strip() in ('', '#N/A', 'None'):
        skipped.append(r)
        continue
    if not ma_kh or str(ma_kh).strip() in ('', '0', 'None'):
        skipped.append(r)
        continue

    # Chuẩn hóa
    ma_kh    = str(ma_kh).strip()
    ten_kh   = str(ten_kh).strip()
    dia_chi  = str(dia_chi).strip() if dia_chi else ''
    mst      = str(mst).strip() if mst else ''
    phone    = str(dien_thoai).strip() if dien_thoai else ''
    mobile   = str(di_dong).strip() if di_dong else ''
    nhom     = str(nhom).strip() if nhom else ''
    cong_no  = cong_no if cong_no else 0
    is_ncc   = 1 if str(la_ncc).strip() == '✓' else 0
    is_kh    = 1  # tất cả đều là khách hàng

    # Xác định loại: company hay person
    # Cá nhân nếu: tên có họ đệm tên (2+ từ không có từ khóa công ty)
    company_keywords = ['công ty', 'cty', 'doanh nghiệp', 'dn ', 'htx', 'tập đoàn',
                        'ngân hàng', 'bệnh viện', 'trường', 'ủy ban', 'văn phòng',
                        'chi nhánh', 'siêu thị', 'cửa hàng', 'shop', 'co.,', 'ltd', 'inc']
    ten_lower = ten_kh.lower()
    company_type = 'company'
    for kw in company_keywords:
        if kw in ten_lower:
            company_type = 'company'
            break
    else:
        # Nếu bắt đầu bằng họ phổ biến → person
        vn_ho = ['nguyễn', 'trần', 'lê', 'phạm', 'huỳnh', 'hoàng', 'phan', 'vũ', 'võ',
                 'đặng', 'bùi', 'đỗ', 'hồ', 'ngô', 'dương', 'lý', 'anh ', 'chị ', 'ông ', 'bà ']
        for ho in vn_ho:
            if ten_lower.startswith(ho):
                company_type = 'person'
                break

    # External ID: thêm prefix để không trùng với module khác
    ext_id = f"misa_{ma_kh}"

    records.append({
        'ext_id': ext_id,
        'ma_kh': ma_kh,
        'ten_kh': ten_kh,
        'company_type': company_type,
        'dia_chi': dia_chi,
        'mst': mst if re.match(r'^\d{10}(\d{3})?$', mst) else '',  # chỉ giữ MST hợp lệ
        'cmnd': str(cmnd).strip() if cmnd else '',
        'phone': phone,
        'mobile': mobile,
        'nhom': nhom,
        'is_kh': is_kh,
        'is_ncc': is_ncc,
        'cong_no': cong_no,
    })

print(f"Đọc đ��ợc: {len(records)} khách hàng hợp lệ")
print(f"Bỏ qua: {len(skipped)} dòng (rỗng/lỗi)")
is_ncc_count = sum(1 for r in records if r['is_ncc'] == 1)
print(f"Trong đó: {len(records)-is_ncc_count} chỉ là KH, {is_ncc_count} vừa là KH vừa là NCC")

# ===== TẠO FILE ODOO IMPORT =====
wb_out = openpyxl.Workbook()
ws1 = wb_out.active
ws1.title = "Import_Odoo"

green  = PatternFill("solid", fgColor="1F7A4A")
yellow = PatternFill("solid", fgColor="FFF2CC")
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

# Header dòng 1: tên field Odoo (dòng này Odoo dùng để map)
odoo_headers = [
    'External ID', 'Name', 'Company Type', 'Street',
    'Phone', 'Mobile', 'VAT',
    'Customer Rank', 'Supplier Rank', 'Notes'
]
for c, h in enumerate(odoo_headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    cell.fill = green
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin

# Header dòng 2: tên tiếng Việt (để bạn đọc dễ hơn)
vn_headers = [
    'Mã định danh', 'Tên KH/NCC', 'Loại (company/person)', 'Địa chỉ',
    'Điện thoại', 'Di động', 'Mã số thuế',
    'Là KH (1/0)', 'Là NCC (1/0)', 'Ghi chú (Công nợ MISA)'
]
for c, h in enumerate(vn_headers, 1):
    cell = ws1.cell(row=2, column=c, value=h)
    cell.font = Font(bold=True, name="Arial", size=9)
    cell.fill = PatternFill("solid", fgColor="DDEEFF")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin

# Dữ liệu từ dòng 3
for r, rec in enumerate(records, 3):
    note = f"MISA: Mã={rec['ma_kh']} | Nhóm={rec['nhom']} | Công nợ={rec['cong_no']:,.0f}"
    if rec['cmnd']:
        note += f" | CMND={rec['cmnd']}"

    row_data = [
        rec['ext_id'],
        rec['ten_kh'],
        rec['company_type'],
        rec['dia_chi'],
        rec['phone'],
        rec['mobile'],
        rec['mst'],
        rec['is_kh'],
        rec['is_ncc'],
        note,
    ]
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.border = thin
        cell.alignment = Alignment(vertical="center")
    ws1.row_dimensions[r].height = 16

# Set column widths
widths = [20, 38, 18, 40, 16, 16, 16, 12, 12, 55]
for c, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

ws1.freeze_panes = "A3"

# ===== SHEET 2: Dữ liệu đầy đủ (lưu để nhập sau) =====
ws2 = wb_out.create_sheet("Du_lieu_day_du_MISA")
full_headers = ['Mã KH', 'Tên KH', 'Địa chỉ', 'Công nợ MISA',
                'Nhóm KH/NCC', 'MST', 'CMND', 'Điện thoại', 'Di động', 'Là NCC']
for c, h in enumerate(full_headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, name="Arial", size=10)
    cell.fill = green
    cell.font = Font(color="FFFFFF", bold=True)
    cell.border = thin

for r, rec in enumerate(records, 2):
    row_data = [
        rec['ma_kh'], rec['ten_kh'], rec['dia_chi'], rec['cong_no'],
        rec['nhom'], rec['mst'], rec['cmnd'], rec['phone'], rec['mobile'],
        '✓' if rec['is_ncc'] else ''
    ]
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.border = thin

out_path = r"D:\odoo\customers_odoo_import.xlsx"
wb_out.save(out_path)
print(f"\nĐã tạo file: {out_path}")
print(f"Sheet 1 'Import_Odoo': {len(records)} dòng → dùng để import v��o Odoo")
print(f"Sheet 2 'Du_lieu_day_du_MISA': lưu đầy đủ cột MISA (công nợ, nhóm, CMND...)")
