import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')

path = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khach_hang.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    # Print first 5 rows
    for r in range(1, min(6, ws.max_row+1)):
        row_data = []
        for c in range(1, ws.max_column+1):
            val = ws.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else "")
        print(f"  Row {r}: {row_data}")
    print(f"  ...")
    # Print last 2 rows
    for r in range(max(6, ws.max_row-1), ws.max_row+1):
        row_data = []
        for c in range(1, ws.max_column+1):
            val = ws.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else "")
        print(f"  Row {r}: {row_data}")
