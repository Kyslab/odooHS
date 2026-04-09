import openpyxl, xmlrpc.client, sys
sys.stdout.reconfigure(encoding='utf-8')

# ---- Kết nối Odoo ----
url = 'http://localhost:8017'
db  = 'odoo_company'
username = 'doanvanky36k21@gmail.com'
password = 'admin'

common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
uid = common.authenticate(db, username, password, {})
if not uid:
    print("Lỗi đăng nhập!"); exit(1)
print(f"Đăng nhập OK (uid={uid})")

models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')

# ---- Đọc file đã chuyển đổi ----
wb = openpyxl.load_workbook(r"D:\odoo\customers_odoo_import.xlsx", data_only=True)
ws = wb['Import_Odoo']

records = []
for r in range(3, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 11)]
    ext_id, name, ctype, street, phone, mobile, vat, is_kh, is_ncc, note = row
    if not name or not ext_id:
        continue
    records.append({
        'ext_id': str(ext_id).strip(),
        'name': str(name).strip(),
        'company_type': str(ctype).strip() if ctype else 'company',
        'street': str(street).strip() if street else '',
        'phone': str(phone).strip() if phone else '',
        'mobile': str(mobile).strip() if mobile else '',
        'vat': str(vat).strip() if vat else '',
        'customer_rank': int(is_kh) if is_kh else 1,
        'supplier_rank': int(is_ncc) if is_ncc else 0,
        'comment': str(note).strip() if note else '',
    })

print(f"Sẽ import {len(records)} bản ghi...")

# ---- Import từng batch 50 records ----
created = 0
updated = 0
errors  = []
BATCH   = 50

for i in range(0, len(records), BATCH):
    batch = records[i:i+BATCH]
    for rec in batch:
        try:
            # Tìm theo external ID
            ext_module = '__import__'
            ext_name   = rec['ext_id']

            existing_ext = models.execute_kw(db, uid, password,
                'ir.model.data', 'search_read',
                [[['module','=', ext_module], ['name','=', ext_name], ['model','=','res.partner']]],
                {'fields': ['res_id'], 'limit': 1})

            vals = {
                'name':          rec['name'],
                'company_type':  rec['company_type'],
                'customer_rank': rec['customer_rank'],
                'supplier_rank': rec['supplier_rank'],
            }
            if rec['street']:  vals['street']  = rec['street']
            if rec['phone']:   vals['phone']   = rec['phone']
            if rec['mobile']:  vals['mobile']  = rec['mobile']
            if rec['vat']:     vals['vat']     = rec['vat']
            if rec['comment']: vals['comment'] = rec['comment']

            if existing_ext:
                # Update
                partner_id = existing_ext[0]['res_id']
                models.execute_kw(db, uid, password, 'res.partner', 'write',
                    [[partner_id], vals])
                updated += 1
            else:
                # Create
                partner_id = models.execute_kw(db, uid, password,
                    'res.partner', 'create', [vals])
                # Lưu external ID
                models.execute_kw(db, uid, password, 'ir.model.data', 'create', [{
                    'module': ext_module,
                    'name':   ext_name,
                    'model':  'res.partner',
                    'res_id': partner_id,
                    'noupdate': True,
                }])
                created += 1

        except Exception as e:
            errors.append(f"  [{rec['ext_id']}] {rec['name']}: {e}")

    done = min(i + BATCH, len(records))
    print(f"  [{done}/{len(records)}] tạo={created} cập nhật={updated} lỗi={len(errors)}")

print(f"\n{'='*50}")
print(f"Hoàn tất!")
print(f"  Tạo mới : {created}")
print(f"  Cập nhật: {updated}")
print(f"  Lỗi     : {len(errors)}")
if errors:
    print("\nChi tiết lỗi:")
    for e in errors[:20]:
        print(e)
