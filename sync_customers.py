"""
sync_customers.py
-----------------
Đọc file Danh_sach_khach_hang.xlsx, phát hiện khách hàng MỚI (chưa có trong Odoo)
và tự động import vào Odoo.

Cách dùng:
  python D:/odoo/sync_customers.py

Chạy tự động: đặt lịch Windows Task Scheduler hoặc chạy thủ công khi cần.
"""

import openpyxl, xmlrpc.client, sys, re, os, json
from datetime import datetime
from odoo_helper import load_all_tags, assign_tag_to_partner
sys.stdout.reconfigure(encoding='utf-8')

# ===================== CẤU HÌNH =====================
EXCEL_FILE   = r"C:\Users\DELL\Downloads\in xong xoa\Danh_sach_khach_hang.xlsx"
SHEET_NAME   = 'Danh sách khách hàng'
HEADER_ROW   = 3   # dòng tiêu đề
DATA_START   = 4   # dòng dữ liệu bắt đầu

ODOO_URL     = 'http://localhost:8017'
ODOO_DB      = 'odoo_company'
ODOO_USER    = 'doanvanky36k21@gmail.com'
ODOO_PASS    = 'admin'

LOG_FILE     = r"D:\odoo\sync_customers.log"
EXT_MODULE   = '__import__'
# ====================================================

def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

def clean(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s in ('None', '#N/A') else s

def detect_company_type(name):
    name_lower = name.lower()
    keywords = ['công ty', 'cty', 'doanh nghiệp', 'dn ', 'htx', 'tập đoàn',
                'ngân hàng', 'bệnh viện', 'trường ', 'ủy ban', 'văn phòng',
                'chi nhánh', 'siêu thị', 'cửa hàng', 'co.,', 'ltd', 'tnhh', 'cổ phần']
    for kw in keywords:
        if kw in name_lower:
            return 'company'
    ho_vn = ['nguyễn', 'trần', 'lê', 'phạm', 'huỳnh', 'hoàng', 'phan', 'vũ', 'võ',
             'đặng', 'bùi', 'đỗ', 'hồ', 'ngô', 'dương', 'lý',
             'anh ', 'chị ', 'ông ', 'bà ', 'chú ', 'cô ']
    for ho in ho_vn:
        if name_lower.startswith(ho):
            return 'person'
    return 'company'

def connect_odoo():
    common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
    if not uid:
        raise Exception("Đăng nhập Odoo thất bại!")
    models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
    return uid, models

def get_tag_cache(models, uid):
    log("Đang load danh mục nhóm (tags)...")
    cache = load_all_tags(models, uid, ODOO_DB, ODOO_PASS)
    log(f"Đã load {len(cache)} nhóm tags")
    return cache

def get_existing_ext_ids(models, uid):
    """Lấy tất cả external ID đã có trong Odoo"""
    records = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
        'ir.model.data', 'search_read',
        [[['module', '=', EXT_MODULE], ['model', '=', 'res.partner']]],
        {'fields': ['name']})
    return {r['name'] for r in records}

def read_excel_customers():
    """Đọc toàn bộ khách hàng từ Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    customers = []
    for r in range(DATA_START, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        stt, ma_kh, ten_kh, dia_chi, cong_no, nhom, mst, phone, cmnd, mobile, la_ncc = row

        ma_kh  = clean(ma_kh)
        ten_kh = clean(ten_kh)

        # Bỏ qua dòng rỗng, lỗi, hoặc dòng tổng
        if not ten_kh or not ma_kh or ma_kh in ('0', 'Tổng'):
            continue
        if ten_kh in ('#N/A',): continue

        mst_clean = clean(mst)
        if not re.match(r'^\d{10}(\d{3})?$', mst_clean):
            mst_clean = ''

        customers.append({
            'ext_id':       f"misa_{re.sub(r'[^a-zA-Z0-9_]', '_', ma_kh)}",
            'ma_kh':        ma_kh,
            'name':         ten_kh,
            'company_type': detect_company_type(ten_kh),
            'street':       clean(dia_chi),
            'phone':        clean(phone),
            'mobile':       clean(mobile),
            'vat':          mst_clean,
            'customer_rank': 1,
            'supplier_rank': 1 if clean(la_ncc) == '✓' else 0,
            'comment':      f"MISA: Mã={ma_kh} | Nhóm={clean(nhom)} | Công nợ={cong_no or 0:,.0f}",
        })
    return customers

def import_to_odoo(models, uid, new_records):
    created, errors = 0, []
    for rec in new_records:
        try:
            vals = {
                'name':          rec['name'],
                'ref':           rec['ma_kh'],   # Mã KH từ MISA
                'company_type':  rec['company_type'],
                'customer_rank': rec['customer_rank'],
                'supplier_rank': rec['supplier_rank'],
            }
            if rec['street']:  vals['street']  = rec['street']
            if rec['phone']:   vals['phone']   = rec['phone']
            if rec['mobile']:  vals['mobile']  = rec['mobile']
            if rec['vat']:     vals['vat']     = rec['vat']
            if rec['comment']: vals['comment'] = rec['comment']

            partner_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'res.partner', 'create', [vals])

            models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                'ir.model.data', 'create', [{
                    'module':   EXT_MODULE,
                    'name':     rec['ext_id'],
                    'model':    'res.partner',
                    'res_id':   partner_id,
                    'noupdate': True,
                }])
            created += 1
            log(f"  ✓ Tạo mới: [{rec['ma_kh']}] {rec['name']}")
            # Gán tag nhóm
            nhom = re.search(r'Nh[oó]m=([A-Z0-9]+)', rec.get('comment',''))
            if nhom:
                assign_tag_to_partner(models, uid, ODOO_DB, ODOO_PASS, partner_id, nhom.group(1), tag_cache)

        except Exception as e:
            errors.append(rec['name'])
            log(f"  ✗ Lỗi [{rec['ma_kh']}] {rec['name']}: {e}")

    return created, errors

# ===================== MAIN =====================
if __name__ == '__main__':
    log("=" * 55)
    log("BẮT ĐẦU ĐỒNG BỘ KHÁCH HÀNG MISA → ODOO")
    log(f"File: {EXCEL_FILE}")

    # 1. Kết nối Odoo
    uid, models = connect_odoo()
    log("Kết nối Odoo OK")
    tag_cache = get_tag_cache(models, uid)

    # 2. Lấy danh sách ext_id đã có
    existing = get_existing_ext_ids(models, uid)
    log(f"Odoo hiện có {len(existing)} bản ghi từ MISA")

    # 3. Đọc Excel
    all_customers = read_excel_customers()
    log(f"Excel có {len(all_customers)} khách hàng hợp lệ")

    # 4. Lọc ra những khách hàng MỚI (chưa có trong Odoo)
    new_customers = [c for c in all_customers if c['ext_id'] not in existing]
    log(f"Phát hiện {len(new_customers)} khách hàng MỚI cần thêm vào Odoo")

    if not new_customers:
        log("Không có gì mới. Kết thúc.")
    else:
        # 5. Import
        created, errors = import_to_odoo(models, uid, new_customers)
        log(f"KẾT QUẢ: Tạo mới={created} | Lỗi={len(errors)}")

    log("ĐỒNG BỘ HOÀN TẤT")
    log("=" * 55)
