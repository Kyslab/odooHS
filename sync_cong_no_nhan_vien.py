# -*- coding: utf-8 -*-
"""
Nhập số dư công nợ đầu kỳ nhân viên từ MISA vào Odoo.

TK 141  - Tạm ứng (Advances to employees)
TK 3341 - Lương phải trả nhân viên

Đồng thời:
- Xóa dòng TK 141 và TK 3341 khỏi bút toán MISC (move_id=2)
- Điều chỉnh TK 4111 trong MISC cho cân bằng
"""

import sys
import psycopg2
import openpyxl
from datetime import date

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE  = r'C:\Users\DELL\Downloads\in xong xoa\Danh_sach_cong_no_nhan_vien.xlsx'
DB_CONN     = dict(host='localhost', port=5432, dbname='odoo_company',
                   user='odoo17', password='odoo17pass')
MOVE_DATE   = date(2024, 12, 31)
MISC_MOVE_ID   = 2
MISC_LINE_141  = 10    # line_id TK141 trong MISC
MISC_LINE_3341 = 23    # line_id TK3341 trong MISC
MISC_LINE_4111 = 27    # line_id TK4111 trong MISC
OLD_4111_CREDIT = 227_394_292_256

ACC_141  = 247   # account_id TK 141
ACC_3341 = 129   # account_id TK 3341
ACC_4111 = 294   # account_id TK 4111
JOURNAL_MISC = 3
COMPANY_ID   = 1

# ── đọc Excel ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active

# map ref → partner_id  (sẽ tra sau)
entries = []   # [(tk, ma_nv, ten, du_no, du_co)]
for r in range(4, ws.max_row + 1):
    stt   = ws.cell(row=r, column=1).value
    if not stt:
        continue
    tk    = str(ws.cell(row=r, column=2).value or '').strip()
    ma    = str(ws.cell(row=r, column=3).value or '').strip()
    ten   = str(ws.cell(row=r, column=4).value or '').strip()
    du_no = float(ws.cell(row=r, column=5).value or 0)
    du_co = float(ws.cell(row=r, column=6).value or 0)
    if du_no == 0 and du_co == 0:
        continue
    entries.append((tk, ma, ten, du_no, du_co))

print(f'Đọc từ Excel: {len(entries)} dòng có số dư')

# ── kết nối DB ────────────────────────────────────────────────────────────────
conn = psycopg2.connect(**DB_CONN)
cur  = conn.cursor()

# Tra partner_id theo ref
cur.execute("SELECT ref, id FROM res_partner WHERE ref ~ '^(LX|NV)[0-9]+$' AND active=TRUE;")
ref_to_pid = {r: pid for r, pid in cur.fetchall()}

# ── xây dựng danh sách journal lines ──────────────────────────────────────────
lines_141  = []   # (partner_id, debit, credit)
lines_3341 = []
not_found  = []

for tk, ma, ten, du_no, du_co in entries:
    pid = ref_to_pid.get(ma)
    if not pid:
        not_found.append((ma, ten))
        continue

    if tk == '141':
        if du_no > 0:
            lines_141.append((pid, du_no, 0.0))        # Dư Nợ → Debit
        elif du_no < 0:
            lines_141.append((pid, 0.0, abs(du_no)))   # Dư Nợ âm → Credit
    elif tk == '3341':
        if du_co > 0:
            lines_3341.append((pid, 0.0, du_co))        # Dư Có → Credit

if not_found:
    print(f'[WARN] Không tìm thấy partner cho: {not_found}')

total_debit_141  = sum(l[1] for l in lines_141)
total_credit_141 = sum(l[2] for l in lines_141)
total_credit_3341 = sum(l[2] for l in lines_3341)

total_debit_non4111  = total_debit_141
total_credit_non4111 = total_credit_141 + total_credit_3341
balance_4111_debit   = total_credit_non4111 - total_debit_non4111

print(f'TK 141  : Debit={total_debit_141:>15,.0f} | Credit={total_credit_141:>15,.0f}  ({len(lines_141)} dòng)')
print(f'TK 3341 : Credit={total_credit_3341:>14,.0f}  ({len(lines_3341)} dòng)')
print(f'TK 4111 : Debit={balance_4111_debit:>15,.0f}  (cân bằng)')

# TK4111 mới trong MISC
total_debit_removed  = total_debit_141   # dư nợ TK141 trong MISC
total_credit_removed = total_credit_3341 + total_credit_141  # dư có TK3341 (và 141 âm) trong MISC
# MISC TK141 có debit=49,063,000; TK3341 có credit=820,685,972
# Thực tế các giá trị trong MISC là tổng net đã bao gồm cả âm
misc_141_debit  = 49_063_000
misc_3341_credit = 820_685_972
new_4111_credit = OLD_4111_CREDIT - misc_3341_credit + misc_141_debit
print(f'\nMISC TK4111 cũ  : {OLD_4111_CREDIT:>20,.0f}')
print(f'MISC TK4111 mới : {new_4111_credit:>20,.0f}')

# ── tạo bút toán mới (move) ───────────────────────────────────────────────────
print('\n--- Tạo bút toán công nợ nhân viên ---')

cur.execute(
    "UPDATE account_move SET state='draft' WHERE id=%s;", (MISC_MOVE_ID,)
)

# Tạo move
cur.execute("""
    INSERT INTO account_move
        (journal_id, date, state, move_type, company_id,
         ref, narration, currency_id, auto_post)
    SELECT %s, %s, 'draft', 'entry', %s,
           'CNĐK-NV', 'Số dư công nợ đầu kỳ nhân viên 31/12/2024',
           (SELECT id FROM res_currency WHERE name='VND' LIMIT 1),
           'no'
    RETURNING id;
""", (JOURNAL_MISC, MOVE_DATE, COMPANY_ID))
new_move_id = cur.fetchone()[0]
print(f'Tạo move_id={new_move_id}')

# Sequence name cho move
cur.execute("""
    UPDATE account_move SET name = 'MISC/' || %s::text WHERE id = %s;
""", (new_move_id, new_move_id))

# ── chèn journal lines ────────────────────────────────────────────────────────
def ins(account_id, partner_id, debit, credit, name):
    cur.execute("""
        INSERT INTO account_move_line
            (move_id, account_id, partner_id, debit, credit,
             name, company_id, currency_id, amount_currency,
             date, date_maturity, display_type)
        SELECT %s, %s, %s, %s, %s,
               %s, %s,
               (SELECT id FROM res_currency WHERE name='VND' LIMIT 1),
               %s - %s,
               %s, %s, 'product'
    """, (new_move_id, account_id, partner_id, debit, credit,
          name, COMPANY_ID, debit, credit, MOVE_DATE, MOVE_DATE))

# TK 141
for pid, db, cr in lines_141:
    ins(ACC_141, pid, db, cr, 'Số dư TK 141 đầu kỳ 2025')
print(f'Đã thêm {len(lines_141)} dòng TK 141')

# TK 3341
for pid, db, cr in lines_3341:
    ins(ACC_3341, pid, db, cr, 'Số dư TK 3341 đầu kỳ 2025')
print(f'Đã thêm {len(lines_3341)} dòng TK 3341')

# TK 4111 cân bằng
ins(ACC_4111, None, balance_4111_debit, 0.0, 'Số dư đối ứng TK 4111')
print(f'Đã thêm dòng TK 4111 (Debit={balance_4111_debit:,.0f})')

# ── post move mới ─────────────────────────────────────────────────────────────
cur.execute("UPDATE account_move SET state='posted' WHERE id=%s;", (new_move_id,))
print(f'Đã post move_id={new_move_id}')

# ── cập nhật MISC (move_id=2): xóa TK141 & TK3341, điều chỉnh TK4111 ─────────
print(f'\n--- Cập nhật MISC (move_id={MISC_MOVE_ID}) ---')

cur.execute("DELETE FROM account_move_line WHERE id = ANY(%s);",
            ([MISC_LINE_141, MISC_LINE_3341],))
print(f'Đã xóa line_id {MISC_LINE_141} (TK141) và {MISC_LINE_3341} (TK3341) khỏi MISC')

cur.execute("UPDATE account_move_line SET credit=%s WHERE id=%s;",
            (new_4111_credit, MISC_LINE_4111))
print(f'TK4111 MISC: {OLD_4111_CREDIT:,.0f} → {new_4111_credit:,.0f}')

cur.execute("UPDATE account_move SET state='posted' WHERE id=%s;", (MISC_MOVE_ID,))

conn.commit()

# ── kiểm tra balance ───────────────────────────────────────────────────────────
cur.execute("""
    SELECT SUM(debit), SUM(credit) FROM account_move_line WHERE move_id=%s;
""", (new_move_id,))
r = cur.fetchone()
print(f'\nKiểm tra move_id={new_move_id}: Debit={r[0]:,.0f} | Credit={r[1]:,.0f} | {"✓ Cân bằng" if abs(r[0]-r[1])<1 else "✗ LỆCH!"}')

cur.execute("""
    SELECT SUM(debit), SUM(credit) FROM account_move_line WHERE move_id=%s;
""", (MISC_MOVE_ID,))
r = cur.fetchone()
print(f'Kiểm tra MISC (id={MISC_MOVE_ID}): Debit={r[0]:,.0f} | Credit={r[1]:,.0f} | {"✓ Cân bằng" if abs(r[0]-r[1])<1 else "✗ LỆCH!"}')

cur.close()
conn.close()
print('\nHoàn tất!')
